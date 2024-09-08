import base64
import io
import json
import os
import random
import re
import sys
import time
import tkinter as tk
from datetime import datetime
from tkinter import ttk

import keyboard
import pygetwindow as gw
import mouse
import openpyxl
import psutil
import pyautogui
import pymsgbox
import pyperclip
import pyttsx4
import requests
import win32clipboard
import win32con
import win32gui
import winsound
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPainter, QPen, QColor, QPixmap
from PyQt5.QtWidgets import QApplication, QWidget
from aip import AipOcr
from dateutil.parser import parse

from functions import get_str_now_time, line_number_increment
from ini控制 import get_ocr_info, get_setting_data_from_ini, extract_resource_folder_path, \
    matched_complete_path_from_resource_folders
from 数据库操作 import (
    get_variable_info,
    set_variable_value,
)
from 网页操作 import WebOption

sys.coinit_flags = 2  # STA
from pywinauto import Application

# from pywinauto.findwindows import ElementNotFoundError

# dic_ = {
#                     'ID': elem_[0],
#                     '图像路径': elem_[1],
#                     '指令类型': elem_[2],
#                     '参数1（键鼠指令）': elem_[3],
#                     '参数2': elem_[4],
#                     '参数3': elem_[5],
#                     '参数4': elem_[6],
#                     '重复次数': elem_[7],
#                     '异常处理': elem_[8]
#                 }

DRIVER = None  # 浏览器驱动


def exit_main_work():
    sys.exit()


def close_browser():
    """关闭浏览器"""
    global DRIVER
    web_option = WebOption()
    web_option.driver = DRIVER
    web_option.close_browser()


def sub_variable(text: str):
    """将text中的变量替换为变量值"""
    new_text = text
    if ("☾" in text) and ("☽" in text):
        variable_dic = get_variable_info("dict")
        for key, value in variable_dic.items():
            new_text = new_text.replace(f"☾{key}☽", str(value))
    return new_text


def get_available_path(image_name_: str, out_mes, is_test=False):
    """组合图片路径，返回可以打开的图片路径，如果路径不存在则重新匹配
    :param is_test: 是否测试
    :param out_mes: 用于输出信息
    :param image_name_: 图片路径或者图片名称
    :return: 可以打开的图片路径，如果仍然不存在则返回None"""

    def search_image_in_folders(image_name_only_, folders):
        for folder_path in folders:
            image_path = os.path.join(folder_path, image_name_only_)
            if os.path.exists(image_path):
                return image_path
        return None

    if os.path.isabs(image_name_):
        if os.path.exists(image_name_):
            return image_name_
        else:
            out_mes.out_mes("原资源文件路径不存在，已重新匹配。", is_test)
            image_name_only = os.path.basename(image_name_)
            res_folder_path = extract_resource_folder_path()
            return search_image_in_folders(image_name_only, res_folder_path)

    else:
        res_folder_path = extract_resource_folder_path()
        return search_image_in_folders(image_name_, res_folder_path)


class TransparentWindow(QWidget):
    """显示框选区域的窗口"""

    def __init__(self):
        """pos(x,y, width, height)"""
        super().__init__()
        # 设置无边框窗口
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Tool)
        self.setWindowOpacity(0.5)  # 设置透明度
        self.setAttribute(Qt.WA_TranslucentBackground)  # 设置背景透明

    def paintEvent(self, event):
        # 绘制边框
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setPen(
            QPen(QColor(255, 0, 0), 5, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin)
        )
        painter.drawRect(self.rect())


class OutputMessage:
    """输出信息，测试时输出到文本框，非测试时输出到主窗口"""

    def __init__(self, command_thread, navigation):
        # 输出的窗口
        self.command_thread = command_thread
        self.navigation = navigation

    def out_mes(self, message: str, is_test: bool = False):
        """输出信息,测试时输出到文本框，非测试时输出到主窗口"""
        if not is_test:
            self.command_thread.show_message(f"----{message}")
        elif is_test:
            self.navigation.textBrowser.append(f"{get_str_now_time()}\t{message}")
        QApplication.processEvents()


def timer(func):
    def func_wrapper(*args, **kwargs):
        from time import time

        time_start = time()
        result = func(*args, **kwargs)
        time_end = time()
        time_spend = time_end - time_start
        print("%s cost time: %.3f s" % (func.__name__, time_spend))
        return result

    return func_wrapper


class ImageClick:
    """图像点击"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        setting_data_dic = get_setting_data_from_ini(
            'Config',
            "持续时间", "时间间隔", "暂停时间"
        )
        self.duration = float(setting_data_dic.get("持续时间"))
        self.interval = float(setting_data_dic.get("时间间隔"))
        self.time_sleep = float(setting_data_dic.get("暂停时间"))
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic = ins_dic  # 指令字典
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数
        :return: 指令参数列表，重复次数"""

        def convert_to_tuple(s):
            """将形如“(x, y)”的字符串转换为元组(x, y)。"""

            def convert_element(element):
                """
                尝试将字符串转换为适当的类型。
                - 如果可以转换为整数，返回整数。
                - 如果可以转换为浮点数，返回浮点数。
                - 否则，返回原字符串。
                """
                try:
                    return int(element)
                except ValueError:
                    try:
                        return float(element)
                    except ValueError:
                        return element

            # 去除字符串中的圆括号和空格
            s = s.strip('()').replace(' ', '')
            # 以逗号分割字符串
            parts = s.split(',')
            if len(parts) == 2:
                # 将每个部分转换为合适的类型并返回元组
                return tuple(map(convert_element, parts))
            else:
                raise ValueError(f"输入格式不正确: {s}")

        # 取重复次数
        re_try = self.ins_dic.get("重复次数")
        # 获取其他参数
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        area_identification = eval(parameter_dic_.get("区域"))  # 是否区域识别
        if area_identification == (0, 0, 0, 0):
            area_identification = None  # 如果没有区域识别则设置为None
        click_map = {
            "左键单击": [1, "left"],
            "左键双击": [2, "left"],
            "右键单击": [1, "right"],
            "右键双击": [2, "right"],
            "仅移动鼠标": [0, "left"],
        }
        list_ins = click_map.get(parameter_dic_.get("动作"))
        # 返回重复次数，点击次数，左键右键，图片名称，是否跳过
        return {
            "重复次数": re_try,
            "点击次数": list_ins[0],
            "左右键": list_ins[1],
            "图像名称": get_available_path(self.ins_dic.get("图像路径"), self.out_mes, self.is_test),
            "异常": parameter_dic_.get("异常"),  # 是否跳过参数
            "灰度": parameter_dic_.get("灰度"),  # 是否灰度识别
            "区域": area_identification,
            "精度": float(parameter_dic_.get("精度", 0.8)),  # 图像匹配精度,
            "点击偏移位置": convert_to_tuple(parameter_dic_.get("点击位置", "(0,0)"))  # 点击偏移位置
        }

    def start_execute(self):
        """开始执行鼠标点击事件"""
        # 解析指令字典
        ins_dic = self.parsing_ins_dic()
        reTry = int(ins_dic.get("重复次数"))
        # 执行图像点击
        for _ in range(reTry):
            self.execute_click(
                click_times=ins_dic.get("点击次数"),
                gray_rec=ins_dic.get("灰度"),
                lOrR=ins_dic.get("左右键"),
                img=ins_dic.get("图像名称"),
                skip=ins_dic.get("异常"),
                area=ins_dic.get("区域"),
                precision=ins_dic.get("精度"),
                click_position=ins_dic.get("点击偏移位置")
            )
            time.sleep(self.time_sleep)

    def execute_click(
            self,
            click_times,
            gray_rec, lOrR, img, skip,
            precision=0.8,
            click_position=(0, 0),
            area=None
    ):
        """执行鼠标点击事件
        :param click_times: 点击次数
        :param gray_rec: 是否灰度识别
        :param lOrR: 左键右键
        :param img: 图像名称
        :param skip: 是否跳过(自动略过、数字)
        :param area: 是否区域识别
        :param precision: 图像匹配精度
        :param click_position: 点击偏移位置"""

        # 4个参数：鼠标点击时间，按钮类型（左键右键中键），图片名称，重复次数
        def image_match_click(location, spend_time, new_click_position):
            if location is not None:
                if not self.is_test:
                    # 参数：位置X，位置Y，点击次数，时间间隔，持续时间，按键
                    self.out_mes.out_mes(
                        f"已找到匹配图片，耗时{spend_time}毫秒。", self.is_test
                    )
                    pyautogui.click(
                        location.x + new_click_position[0],
                        location.y + new_click_position[1],
                        clicks=click_times,
                        interval=self.interval,
                        duration=self.duration,
                        button=lOrR,
                    )
                elif self.is_test:
                    self.out_mes.out_mes(
                        f"已找到匹配图片，耗时{spend_time}毫秒。", self.is_test
                    )
                    # 移动鼠标到图片位置
                    pyautogui.moveTo(
                        location.x + new_click_position[0],
                        location.y + new_click_position[1],
                        duration=0.2
                    )

        def get_new_click_position(old_click_position):
            """获取随机点击位置"""
            if old_click_position == ('随机', '随机'):
                pixmap = QPixmap(img)
                pixmap_width = pixmap.width()
                pixmap_height = pixmap.height()
                # 随机生成点击位置
                new_click_position = (
                    random.randint(-pixmap_width // 2, pixmap_width // 2),
                    random.randint(-pixmap_height // 2, pixmap_height // 2)
                )
                return new_click_position
            else:
                return old_click_position

        min_search_time = 1 if skip == "自动略过" else float(skip)
        is_skip = True if skip == "自动略过" else False
        new_click_position_ = get_new_click_position(click_position)
        try:
            # 显示信息
            self.out_mes.out_mes(f"正在查找匹配图像...", self.is_test)
            QApplication.processEvents()
            # 记录开始时间
            start_time = time.time()
            location_ = pyautogui.locateCenterOnScreen(
                image=img,
                confidence=precision,
                minSearchTime=min_search_time,
                grayscale=gray_rec,
                region=area,
            )
            if location_:  # 如果找到图像
                spend_time_ = int((time.time() - start_time) * 1000)  # 计算耗时
                image_match_click(location_, spend_time_, new_click_position_)
            elif not location_:  # 如果未找到图像
                self.out_mes.out_mes("未找到匹配图像", self.is_test)
                raise FileNotFoundError
        except OSError:
            self.out_mes.out_mes(
                "文件下未找到png图像，请检查文件是否存在！", self.is_test
            )
            raise FileNotFoundError
        except TypeError:
            self.out_mes.out_mes("图像路径不存在！", self.is_test)
            raise TypeError
        except pyautogui.ImageNotFoundException:
            if not is_skip:
                self.out_mes.out_mes("未找到匹配图像", self.is_test)
                raise FileNotFoundError
            else:
                self.out_mes.out_mes("未找到匹配图像，已自动略过", self.is_test)


class MultipleImagesClick:
    """多图点击"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        setting_data_dic = get_setting_data_from_ini(
            'Config',
            "持续时间", "时间间隔", "暂停时间"
        )
        self.duration = float(setting_data_dic.get("持续时间"))
        self.interval = float(setting_data_dic.get("时间间隔"))
        self.time_sleep = float(setting_data_dic.get("暂停时间"))
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        re_try = self.ins_dic.get('重复次数')
        img_name_list = str(self.ins_dic.get('图像路径')).split('、')
        img_path_list = [matched_complete_path_from_resource_folders(img_name) for img_name in img_name_list]
        parameter_dic_ = eval(self.ins_dic.get('参数1（键鼠指令）'))
        area_identification = None if eval(parameter_dic_.get("区域")) == (0, 0, 0, 0) \
            else eval(parameter_dic_.get("区域"))
        # 返回参数字典
        return {
            "重复次数": re_try,
            "图像列表": img_path_list,
            "灰度": parameter_dic_.get("灰度"),  # 是否灰度识别
            "区域": area_identification,
            "动作": parameter_dic_.get("动作"),  # 动作
            "异常": parameter_dic_.get("异常"),  # 是否自动跳过
            "精度": float(parameter_dic_.get("精度", 0.8)),  # 图像匹配精度
        }

    def start_execute(self):
        """开始执行鼠标点击事件"""
        # 解析指令字典
        ins_dic = self.parsing_ins_dic()
        click_map = {
            "左键单击": [1, "left"],
            "左键双击": [2, "left"],
            "右键单击": [1, "right"],
            "右键双击": [2, "right"],
            "仅移动鼠标": [0, "left"],
        }
        # 执行图像点击
        reTry = ins_dic.get("重复次数")
        for _ in range(reTry):
            self.execute_click(
                click_times=click_map.get(ins_dic.get("动作"))[0],
                gray_rec=ins_dic.get("灰度"),
                lOrR=click_map.get(ins_dic.get("动作"))[1],
                img_path_list=ins_dic.get("图像列表"),
                area=ins_dic.get("区域"),
                skip=ins_dic.get("异常"),
                precision=ins_dic.get("精度"),
            )
            if reTry > 1:
                time.sleep(self.time_sleep)

    def execute_click(self, click_times, gray_rec, lOrR, img_path_list, skip, precision, area=None):
        """执行鼠标点击事件
        :param click_times: 点击次数
        :param gray_rec: 是否灰度识别
        :param lOrR: 左键右键
        :param img_path_list: 图像路径列表
        :param skip: 是否跳过
        :param area: 是否区域识别
        :param precision: 图像匹配精度"""

        # 4个参数：鼠标点击时间，按钮类型（左键右键中键），图片名称，重复次数
        def image_match_click(location, spend_time, img_name):
            if location is not None:
                if not self.is_test:
                    # 参数：位置X，位置Y，点击次数，时间间隔，持续时间，按键
                    self.out_mes.out_mes(
                        f"已找到{img_name}，耗时{spend_time}毫秒。", self.is_test
                    )
                    pyautogui.click(
                        location.x,
                        location.y,
                        clicks=click_times,
                        interval=self.interval,
                        duration=self.duration,
                        button=lOrR,
                    )
                elif self.is_test:
                    self.out_mes.out_mes(
                        f"已找到匹配图片，耗时{spend_time}毫秒。", self.is_test
                    )
                    # 移动鼠标到图片位置
                    pyautogui.moveTo(location.x, location.y, duration=0.2)

        for img in img_path_list:
            # 显示信息
            img_name_ = os.path.basename(img)
            self.out_mes.out_mes(f"正在查找匹配图像 {img_name_}...", self.is_test)
            QApplication.processEvents()
            # 记录开始时间
            start_time = time.time()
            try:
                location_ = pyautogui.locateCenterOnScreen(
                    image=img,
                    confidence=precision,
                    grayscale=gray_rec,
                    minSearchTime=0,
                    region=area,
                )
                if location_:  # 如果找到图像
                    spend_time_ = int((time.time() - start_time) * 1000)  # 计算耗时
                    image_match_click(location_, spend_time_, img_name_)
                    return  # 一旦找到其中一个图像并点击，直接返回
            except pyautogui.ImageNotFoundException:
                pass
            except OSError or TypeError:
                self.out_mes.out_mes(
                    f'本地文件"{img}"不存在，请检查文件是否存在！', self.is_test
                )
                if skip != "自动略过":
                    raise FileNotFoundError(f'本地文件"{img}"不存在！')

        # 如果所有图像都未找到
        self.out_mes.out_mes("未找到全部的匹配图像", self.is_test)
        if skip != "自动略过":
            raise FileNotFoundError("未找到全部的匹配图像")


class CoordinateClick:
    """坐标点击"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        setting_data_dic = get_setting_data_from_ini(
            'Config',
            "持续时间", "时间间隔", "暂停时间"
        )
        self.duration = float(setting_data_dic.get("持续时间"))
        self.interval = float(setting_data_dic.get("时间间隔"))
        self.time_sleep = float(setting_data_dic.get("暂停时间"))
        self.out_mes = outputmessage  # 用于输出信息
        self.ins_dic = ins_dic  # 指令字典
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        re_try = self.ins_dic.get("重复次数")
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        # 取x,y坐标的值
        x_ = int(parameter_dic_.get("坐标").split("-")[0])
        y_ = int(parameter_dic_.get("坐标").split("-")[1])
        z_ = int(parameter_dic_.get("自定义次数"))
        click_map = {
            "左键单击": [1, "left", x_, y_],
            "左键双击": [2, "left", x_, y_],
            "右键单击": [1, "right", x_, y_],
            "右键双击": [2, "right", x_, y_],
            "左键（自定义次数）": [z_, "left", x_, y_],
            "仅移动鼠标": [0, "left", x_, y_],
        }
        list_ins = click_map.get(parameter_dic_.get("动作"))
        # 返回重复次数，点击次数，左键右键，x坐标，y坐标
        return re_try, list_ins[0], list_ins[1], list_ins[2], list_ins[3]

    def start_execute(self):
        """开始执行鼠标点击事件"""
        # 获取参数
        reTry, click_times, lOrR, x__, y__ = self.parsing_ins_dic()
        # 执行坐标点击
        for _ in range(reTry):
            self.coor_click(click_times, lOrR, x__, y__)
            time.sleep(self.time_sleep)

    def coor_click(self, click_times, lOrR, x__, y__):
        pyautogui.click(
            x=x__,
            y=y__,
            clicks=click_times,
            interval=self.interval,
            duration=self.duration,
            button=lOrR,
        )
        if click_times == 0:
            self.out_mes.out_mes(f"移动鼠标到{x__}-{y__}", self.is_test)
        else:
            self.out_mes.out_mes(f"执行坐标{x__}-{y__}点击", self.is_test)


class TimeWaiting:
    """等待"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 执行线程
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def start_execute(self):
        """从指令字典中解析出指令参数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        wait_type = parameter_dic_.get("类型")
        if wait_type == "时间等待":
            wait_time = int(parameter_dic_.get("时长"))
            unit = parameter_dic_.get("单位")
            self.out_mes.out_mes(f"等待时长{wait_time}{unit}", self.is_test)
            self.stop_time(wait_time, unit)
        elif wait_type == "定时等待":
            target_time = parameter_dic_.get("时间")
            interval_time = int(parameter_dic_.get("检测频率"))
            # 检查目标时间是否大于当前时间
            # if parse(target_time) > datetime.now():
            self.wait_to_time(target_time, interval_time)
        elif wait_type == "随机等待":
            min_time, min_unit = parameter_dic_.get("最小").split("-")
            max_time, max_unit = parameter_dic_.get("最大").split("-")
            if min_unit != max_unit:  # 如果单位不一致则统一单位
                # 统一单位，转换为毫秒，生成随机等待时间
                min_time_u = self.unified_unit(int(min_time), min_unit)
                max_time_u = self.unified_unit(int(max_time), max_unit)
                wait_time = random.randint(min_time_u, max_time_u)
                self.out_mes.out_mes(f"随机等待时间{wait_time}毫秒", self.is_test)
                self.stop_time(wait_time, "毫秒")
            elif min_unit == max_unit:
                wait_time = random.randint(int(min_time), int(max_time))
                self.out_mes.out_mes(f"随机等待时间{wait_time}{min_unit}", self.is_test)
                self.stop_time(wait_time, min_unit)

    @staticmethod
    def unified_unit(time_, unit_):
        """统一单位，转换为毫秒"""
        if unit_ == "秒":
            return time_ * 1000
        elif unit_ == "分钟":
            return time_ * 1000 * 60
        elif unit_ == "毫秒":
            return time_

    def wait_to_time(self, target_time, interval):
        """检查时间，指定时间则执行操作
        :param target_time: 目标时间
        :param interval: 时间间隔"""
        sleep_time = int(interval) / 1000
        show_times = 1  # 显示时间的间隔

        while True:
            now = datetime.now()
            if show_times == 1:
                self.out_mes.out_mes(
                    "当前为：%s" % now.strftime("%H:%M:%S"), self.is_test
                )
                self.out_mes.out_mes("等待至：%s" % target_time, self.is_test)
                self.out_mes.out_mes("等待中......", self.is_test)
                show_times = sleep_time
            if now >= parse(target_time):
                self.out_mes.out_mes("时间已到，退出等待", self.is_test)
                break
            # 时间暂停
            time.sleep(sleep_time)
            show_times += sleep_time

    def stop_time(self, seconds, uint):
        """暂停时间
        :param seconds: 暂停时间
        :param uint: 时间单位（秒、分钟、毫秒）"""

        def wait_time(seconds_):
            for i_ in range(seconds_):
                # 显示剩下等待时间
                self.out_mes.out_mes(
                    "等待中...剩余%d秒" % (seconds_ - i_), self.is_test
                )
                time.sleep(1)

        if uint == "秒":
            wait_time(seconds)
        elif uint == "分钟":
            wait_time(seconds * 60)
        elif uint == "毫秒":
            self.out_mes.out_mes("等待中...共%d毫秒" % seconds, self.is_test)
            time.sleep(seconds / 1000)


class ImageWaiting:
    """图片等待"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number

    def wait_to_image(self, image, wait_instruction_type, timeout_period, confidence, region=None):
        """执行图片等待"""
        if wait_instruction_type == "等待到指定图像出现":
            self.out_mes.out_mes("正在等待指定图像出现中...", self.is_test)
            QApplication.processEvents()
            location = pyautogui.locateCenterOnScreen(
                image=image, confidence=confidence, minSearchTime=timeout_period, region=region
            )
            if location:
                self.out_mes.out_mes("目标图像已经出现，等待结束", self.is_test)
                QApplication.processEvents()
        elif wait_instruction_type == "等待到指定图像消失":
            vanish = True
            while vanish:
                try:
                    pyautogui.locateCenterOnScreen(
                        image=image, confidence=confidence, minSearchTime=1, region=region
                    )
                except pyautogui.ImageNotFoundException:
                    self.out_mes.out_mes("目标图像已经消失，等待结束", self.is_test)
                    QApplication.processEvents()
                    vanish = False
                else:
                    time.sleep(0.5)

    def start_execute(self):
        """执行图片等待"""
        image_path = self.ins_dic.get("图像路径")
        # 获取其他参数
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        wait_instruction_type = parameter_dic_.get("等待类型")
        timeout_period = int(parameter_dic_.get("超时时间"))
        confidence = float(parameter_dic_.get("精度", 0.8))
        area_identification = (
            None
            if eval(parameter_dic_.get("区域")) == (0, 0, 0, 0)
            else eval(parameter_dic_.get("区域"))
        )
        self.wait_to_image(
            image_path, wait_instruction_type, timeout_period, confidence, area_identification
        )


class RollerSlide:
    """滑动鼠标滚轮"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = float(get_setting_data_from_ini("Config", "暂停时间"))
        self.out_mes = outputmessage  # 用于输出信息
        self.ins_dic = ins_dic  # 指令字典
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    @staticmethod
    def parsing_ins_dic(parameter_dic_):
        """解析指令字典"""
        type_ = parameter_dic_.get("类型")
        if type_ == "滚轮滑动":
            scroll_direction = str(parameter_dic_.get("方向"))
            scroll_distance_ = int(parameter_dic_.get("距离"))
            scroll_distance = (
                scroll_distance_ if scroll_direction == "↑" else -scroll_distance_
            )
            return scroll_direction, scroll_distance
        elif type_ == "随机滚轮滑动":
            min_distance = int(parameter_dic_.get("最小距离"))
            max_distance = int(parameter_dic_.get("最大距离"))
            scroll_direction = random.choice(["↑", "↓"])
            scroll_distance_ = random.randint(min_distance, max_distance)
            scroll_distance = (
                scroll_distance_ if scroll_direction == "↑" else -scroll_distance_
            )
            return scroll_direction, scroll_distance

    def start_execute(self):
        """执行重复次数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        type_ = parameter_dic_.get("类型")
        re_try = self.ins_dic.get("重复次数")
        scroll_direction, scroll_distance = self.parsing_ins_dic(parameter_dic_)
        # 执行滚轮滑动
        if re_try == 1:
            self.wheel_slip(scroll_direction, scroll_distance, type_)
        elif re_try > 1:
            i = 1
            while i < re_try + 1:
                self.wheel_slip(scroll_direction, scroll_distance, type_)
                i += 1
                time.sleep(self.time_sleep)

    def wheel_slip(self, scroll_direction, scroll_distance, type_):
        """滚轮滑动事件"""
        pyautogui.scroll(scroll_distance)
        self.out_mes.out_mes(
            f"{type_}{scroll_direction}{scroll_distance}距离", self.is_test
        )


class TextInput:
    """输入文本"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        setting_data_dic = get_setting_data_from_ini(
            'Config',
            "时间间隔", "暂停时间")
        self.interval = float(setting_data_dic.get("时间间隔"))
        self.time_sleep = float(setting_data_dic.get("暂停时间"))
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number

    def start_execute(self):
        """解析指令字典"""
        input_value = sub_variable(self.ins_dic.get("图像路径"))
        special_control_judgment = eval(
            eval(self.ins_dic.get("参数1（键鼠指令）")).get("手动输入")
        )
        # 执行文本输入
        self.text_input(input_value, special_control_judgment)

    def text_input(self, input_value, special_control_judgment):
        """文本输入事件
        :param input_value: 输入的文本
        :param special_control_judgment: 是否为特殊控件"""
        value_str = str(input_value)
        if not special_control_judgment:
            pyperclip.copy(value_str)
            pyautogui.hotkey("ctrl", "v")
            time.sleep(self.time_sleep)
            self.out_mes.out_mes("执行文本输入：%s" % value_str, self.is_test)
        elif special_control_judgment:
            pyautogui.typewrite(value_str, interval=self.interval)
            self.out_mes.out_mes("执行模拟手动文本输入：%s" % value_str, self.is_test)
            time.sleep(self.time_sleep)


class MoveMouse:
    """移动鼠标"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        setting_data_dic = get_setting_data_from_ini(
            'Config',
            "持续时间", "暂停时间")
        self.duration = float(setting_data_dic.get("持续时间"))
        self.time_sleep = float(setting_data_dic.get("暂停时间"))
        self.out_mes = outputmessage  # 用于输出信息
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def start_execute(self):
        """执行重复次数"""
        re_try = self.ins_dic.get("重复次数")
        # 执行滚轮滑动
        for _ in range(re_try):
            self.mouse_move_fun()  # 执行鼠标移动
            time.sleep(self.time_sleep)

    def mouse_move_fun(self) -> None:
        """执行鼠标移动"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        type_ = parameter_dic_.get("类型")
        if type_ == "直线移动":
            self.mouse_moves(parameter_dic_.get("方向"), parameter_dic_.get("距离"))
        elif type_ == "随机移动":
            random_type = parameter_dic_.get("随机")
            if random_type == "类型1":
                self.mouse_moves_random_1()
            elif random_type == "类型2":
                self.mouse_moves_random_2()
        elif type_ == "指定坐标":
            x = int(parameter_dic_.get("坐标").split(",")[0])
            y = int(parameter_dic_.get("坐标").split(",")[1])
            duration = float(parameter_dic_.get("持续"))
            self.move_mouse_to_coordinates(x, y, duration)
        elif type_ == "变量坐标":
            var_name = parameter_dic_.get("变量")
            self.variable_coordinates(var_name)

    def mouse_moves(self, direction, distance):
        """鼠标移动事件"""
        # 相对于当前位置移动鼠标
        directions = {"↑": (0, -1), "↓": (0, 1), "←": (-1, 0), "→": (1, 0)}
        if direction in directions:
            x, y = directions.get(direction)
            pyautogui.moveRel(
                x * int(distance), y * int(distance), duration=self.duration
            )
        self.out_mes.out_mes(
            "直线移动鼠标%s%s像素距离" % (direction, distance), self.is_test
        )

    def mouse_moves_random_1(self):
        """鼠标移动事件"""
        screen_width, screen_height = pyautogui.size()
        # 随机生成坐标
        x = random.randint(0, screen_width)
        y = random.randint(0, screen_height)
        # 随机生成时间
        duration_ran = random.uniform(0.1, 0.9)
        try:
            pyautogui.moveTo(x, y, duration=duration_ran)
            self.out_mes.out_mes("随机移动鼠标", self.is_test)
        except pyautogui.FailSafeException:
            pass

    def mouse_moves_random_2(self):
        """鼠标移动事件"""
        directions = {"↑": (0, -1), "↓": (0, 1), "←": (-1, 0), "→": (1, 0)}
        direction = random.choice(list(directions.keys()))
        if direction in directions:
            x, y = directions.get(direction)
            distance = random.randint(1, 500)
            duration_ran = random.uniform(0.1, 0.9)
            try:
                pyautogui.moveRel(x * distance, y * distance, duration=duration_ran)
                self.out_mes.out_mes("随机移动鼠标", self.is_test)
            except pyautogui.FailSafeException:
                pass

    def move_mouse_to_coordinates(self, x: int, y: int, duration: float):
        """移动鼠标到指定坐标"""
        pyautogui.moveTo(x, y, duration=duration)
        self.out_mes.out_mes(
            f"移动鼠标到{x}:{y}，持续{round(duration, 2)}秒", self.is_test
        )

    def variable_coordinates(self, var_name):
        """变量坐标"""
        var_value = get_variable_info("dict").get(var_name)
        try:
            x, y = var_value.split(",")
            pyautogui.moveTo(int(x), int(y), duration=0)
            self.out_mes.out_mes(
                f"移动鼠标到变量：{var_name}，值为{x, y}", self.is_test
            )
        except Exception as e:
            print("移动鼠标到变量位置", e)
            self.out_mes.out_mes("变量坐标值不符合坐标格式，移动失败！", self.is_test)


class PressKeyboard:
    """模拟按下键盘"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = float(get_setting_data_from_ini(
            'Config',
            "暂停时间"))
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number

    def parsing_ins_dic(self):
        """解析指令字典"""
        re_try = self.ins_dic.get("重复次数")
        parameter_dic = eval(self.ins_dic.get("参数1（键鼠指令）"))
        keys = parameter_dic.get("按键")
        duration = float(parameter_dic.get("按压时长"))
        return re_try, keys, duration

    def start_execute(self):
        """执行重复次数"""
        re_try, keys, duration = self.parsing_ins_dic()
        for _ in range(re_try):
            self.press_keyboard(keys, duration)
            time.sleep(self.time_sleep)

    def press_keyboard(self, key, duration):
        """键盘按键事件
        :param key: 按键
        :param duration: 按键持续时间(毫秒)"""
        keys = key.split('+')
        # 按下组合键
        if len(keys) > 1:
            keyboard.press_and_release('+'.join(keys))
            self.out_mes.out_mes(f"按下组合键：{key}", self.is_test)
        else:
            key_ = keys[0].lower()
            keyboard.press(key_)
            time.sleep(duration / 1000)
            keyboard.release(key_)
            self.out_mes.out_mes(f"按下按键：{key}，持续{duration}毫秒", self.is_test)


class MiddleActivation:
    """鼠标中键激活"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = float(get_setting_data_from_ini('Config', "暂停时间"))
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number

    def start_execute(self):
        """执行重复次数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        command_type = parameter_dic_.get("类型")
        click_count = int(parameter_dic_.get("次数", 1))
        re_try = self.ins_dic.get("重复次数")
        # 执行滚轮滑动
        for _ in range(re_try):
            self.middle_mouse_button(command_type, click_count)
            time.sleep(self.time_sleep)

    def middle_mouse_button(self, command_type, click_times):
        """中键点击事件"""
        self.out_mes.out_mes("等待按下鼠标中键中...按下F11键退出", self.is_test)
        QApplication.processEvents()
        mouse.wait(button="middle")
        try:
            if command_type == "模拟点击":
                self.simulated_mouse_click(click_times, "左键")
                self.out_mes.out_mes(f"执行鼠标点击{click_times}次", self.is_test)
            elif command_type == "结束等待":
                pass
        except OSError:
            # 弹出提示框。提示检查鼠标是否连接
            self.out_mes.out_mes("连接失败，请检查鼠标是否连接正确。", self.is_test)

    @staticmethod
    def simulated_mouse_click(click_times, lOrR):
        """模拟鼠标点击
        :param click_times: 点击次数
        :param lOrR: (左键、右键)"""
        button = "left" if lOrR == "左键" else "right"
        for i in range(click_times):
            mouse.click(button=button)


class MouseClick:
    """鼠标在当前位置点击"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = float(get_setting_data_from_ini("Config", "暂停时间"))
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number

    def parsing_ins_dic(self):
        """解析指令字典"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        button_type = parameter_dic_.get("鼠标")
        click_times = int(parameter_dic_.get("次数"))
        duration = float(parameter_dic_.get("按压"))
        interval = float(parameter_dic_.get("间隔"))
        auxiliary = parameter_dic_.get("辅助键", "")
        return button_type, click_times, duration, interval, auxiliary

    def start_execute(self):
        """Execute repetitions"""
        button_type, click_times, duration, interval, auxiliary = self.parsing_ins_dic()
        re_try = self.ins_dic.get("重复次数")
        for _ in range(re_try):
            self.simulated_mouse_click(
                click_times, button_type, duration, interval, auxiliary
            )
            time.sleep(self.time_sleep)

    def simulated_mouse_click(self, click_times, lOrR, duration, interval, auxiliary):
        """模拟鼠标点击
        :param duration: 按压时长，单位：毫秒
        :param interval: 时间间隔，单位：毫秒
        :param click_times: 点击次数
        :param lOrR: (左键、右键)
        :param auxiliary: 辅助键，默认为空"""
        button = "left" if lOrR == "左键" else "right"
        if auxiliary:
            auxiliary = auxiliary.lower()
        for i in range(click_times):
            if auxiliary:
                keyboard.press(auxiliary)  # 按下辅助键
            mouse.press(button=button)
            time.sleep(duration / 1000)
            mouse.release(button=button)
            if auxiliary:
                keyboard.release(auxiliary)  # 释放辅助键
            time.sleep(interval / 1000)
        self.out_mes.out_mes(f"鼠标在当前位置点击{click_times}次", self.is_test)


class InformationEntry:
    """从Excel中录入信息到窗口"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = float(get_setting_data_from_ini("Config", "暂停时间"))
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number
        # 图像点击、文本输入的部分功能
        self.image_click = ImageClick(self.out_mes, self.ins_dic)
        self.text_input = TextInput(self.out_mes, self.ins_dic)

    def parsing_ins_dic(self):
        """解析指令字典"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        image_path = get_available_path(
            self.ins_dic.get("图像路径"), self.out_mes, self.is_test
        )
        excel_path = get_available_path(
            parameter_dic_.get("工作簿"), self.out_mes, self.is_test
        )
        list_dic = {
            "点击次数": 3,
            "按钮类型": "left",
            "工作簿路径": excel_path,
            "工作表名称": parameter_dic_.get("工作表"),
            "图像路径": image_path,
            "单元格位置": parameter_dic_.get("单元格"),
            "行号递增": eval(parameter_dic_.get("递增")),
            "特殊控件输入": eval(parameter_dic_.get("模拟输入")),
            "超时报错": parameter_dic_.get("异常"),
            "异常处理": self.ins_dic.get("异常处理"),
        }
        return list_dic

    def start_execute(self):
        """执行重复次数"""
        re_try = self.ins_dic.get("重复次数")
        # 执行滚轮滑动
        if re_try >= 1:
            for _ in range(re_try):
                self.information_entry()
                time.sleep(self.time_sleep)

    def information_entry(self):
        """信息录入"""
        list_dic = self.parsing_ins_dic()
        # 获取excel表格中的值
        cell_value = self.extra_excel_cell_value(
            list_dic.get("工作簿路径"),
            list_dic.get("工作表名称"),
            list_dic.get("单元格位置"),
            list_dic.get("行号递增"),
            self.cycle_number,
        )
        self.image_click.is_test = self.is_test
        self.image_click.execute_click(
            click_times=list_dic.get("点击次数"),
            gray_rec=False,
            lOrR=list_dic.get("按钮类型"),
            img=list_dic.get("图像路径"),
            skip=list_dic.get("超时报错"),
        )
        self.text_input.is_test = self.is_test
        self.text_input.text_input(cell_value, list_dic.get("特殊控件输入"))
        if self.is_test:
            self.out_mes.out_mes("测试已完成！", self.is_test)
        else:
            self.out_mes.out_mes("已执行信息录入", self.is_test)

    def extra_excel_cell_value(
            self, excel_path, sheet_name, cell_position, line_number_increment_, number
    ):
        """获取excel表格中的值
        :param excel_path: excel表格路径
        :param sheet_name: 表格名称
        :param cell_position: 单元格位置
        :param line_number_increment_: 行号递增
        :param number: 循环次数"""
        cell_value = None
        try:
            # 打开excel表格
            wb = openpyxl.load_workbook(excel_path)
            # 选择表格
            sheet = wb[str(sheet_name)]
            if not line_number_increment_:
                # 获取单元格的值
                cell_value = sheet[cell_position].value
                self.out_mes.out_mes(
                    f"获取到的单元格值为：{str(cell_value)}", self.is_test
                )
            elif line_number_increment_:
                # 获取行号递增的单元格的值
                column_number = re.findall(r"[a-zA-Z]+", cell_position)[0]
                line_number = (
                        int(re.findall(r"\d+\.?\d*", cell_position)[0]) + number - 1
                )
                new_cell_position = column_number + str(line_number)
                cell_value = sheet[new_cell_position].value
                self.out_mes.out_mes(
                    f"获取到的单元格值为：{str(cell_value)}", self.is_test
                )
            return cell_value
        except FileNotFoundError:
            self.out_mes.out_mes("没有找到工作簿", self.is_test)
            raise FileNotFoundError("没有找到工作簿")
        except KeyError:
            self.out_mes.out_mes("没有找到工作表", self.is_test)
            raise FileNotFoundError("没有找到工作表")
        except AttributeError:
            self.out_mes.out_mes("没有找到单元格", self.is_test)
            raise FileNotFoundError("没有找到单元格")


class OpenWeb:
    """打开网页"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        self.out_mes = outputmessage  # 主窗口
        self.ins_dic = ins_dic  # 指令字典
        # 网页控制的部分功能
        self.web_option = WebOption(self.out_mes)
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def start_execute(self):
        """执行重复次数"""
        url = self.ins_dic.get("图像路径")
        self.out_mes.out_mes("正在打开网页...等待中...", self.is_test)
        global DRIVER
        DRIVER = self.web_option.open_driver(url, True)
        self.out_mes.out_mes("已打开网页", self.is_test)


class EleControl:
    """网页控制"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        self.out_mes = outputmessage  # 主窗口
        self.ins_dic = ins_dic  # 指令字典
        # 网页控制的部分功能
        self.web_option = WebOption(self.out_mes)
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """解析指令字典"""
        image_path = self.ins_dic.get("图像路径")
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        list_dic = {
            "元素类型": parameter_dic_.get("元素类型"),
            "元素值": image_path,
            "操作类型": parameter_dic_.get("操作"),
            "文本内容": sub_variable(parameter_dic_.get("文本")),
            "超时类型": parameter_dic_.get("超时类型"),
        }
        return list_dic

    def start_execute(self):
        """执行重复次数"""
        list_ins_ = self.parsing_ins_dic()
        # 执行网页操作
        global DRIVER
        self.web_option.driver = DRIVER
        self.web_option.text = list_ins_.get("文本内容")
        self.web_option.single_shot_operation(
            action=list_ins_.get("操作类型"),
            element_value_=list_ins_.get("元素值"),
            element_type_=list_ins_.get("元素类型"),
            timeout_type_=list_ins_.get("超时类型"),
        )
        self.out_mes.out_mes("已执行元素控制", self.is_test)


class WebEntry:
    """将Excel中的值录入网页"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.InformationEntry = InformationEntry(self.out_mes, self.ins_dic)
        # 网页控制的部分功能
        self.web_option = WebOption(self.out_mes)
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """解析指令字典"""
        image_path = get_available_path(self.ins_dic.get("图像路径"), self.out_mes, self.is_test)
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        list_dic = {
            "工作簿路径": image_path,
            "工作表名称": parameter_dic_.get("工作表"),
            "元素类型": parameter_dic_.get("元素类型"),
            "元素值": parameter_dic_.get("元素值"),
            "单元格位置": parameter_dic_.get("单元格"),
            "行号递增": eval(parameter_dic_.get("行号递增")),
            "超时类型": parameter_dic_.get("超时类型"),
        }
        return list_dic

    def start_execute(self):
        """执行重复次数"""
        list_ins_ = self.parsing_ins_dic()
        # 获取excel表格中的值
        cell_value = self.InformationEntry.extra_excel_cell_value(
            list_ins_.get("工作簿路径"),
            list_ins_.get("工作表名称"),
            list_ins_.get("单元格位置"),
            list_ins_.get("行号递增"),
            self.cycle_number,
        )
        # 执行网页操作
        global DRIVER
        self.web_option.driver = DRIVER
        self.web_option.text = cell_value
        self.out_mes.out_mes("已获取到单元格值", self.is_test)
        self.web_option.single_shot_operation(
            action="输入内容",
            element_value_=list_ins_.get("元素值"),
            element_type_=list_ins_.get("元素类型"),
            timeout_type_=list_ins_.get("超时类型"),
        )
        self.out_mes.out_mes("已执行信息录入", self.is_test)


class MouseDrag:
    """鼠标拖拽"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = float(get_setting_data_from_ini("Config", "暂停时间"))
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number

    def parsing_ins_dic(self):
        """解析指令字典"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        start_position = tuple(map(int, parameter_dic_.get("开始位置").split(",")))
        stop_position = tuple(map(int, parameter_dic_.get("结束位置").split(",")))
        duration_ = float(parameter_dic_.get("移动速度"))
        start_random = eval(parameter_dic_.get("开始随机"))
        stop_random = eval(parameter_dic_.get("结束随机"))
        return start_position, stop_position, start_random, stop_random, duration_

    def mouse_drag(self, start_position, end_position, duration_=300):
        """鼠标拖拽事件"""
        pyautogui.moveTo(start_position[0], start_position[1], duration=duration_//1000)
        pyautogui.dragTo(end_position[0], end_position[1], duration=duration_//1000)
        self.out_mes.out_mes(
            "鼠标拖拽%s到%s" % (str(start_position), str(end_position)), self.is_test
        )

    @staticmethod
    def random_position(position, random_range=100):
        """设置随机坐标"""
        if random_range == 0:
            return position
        x, y = position
        x_random = random.randint(-random_range, random_range)
        y_random = random.randint(-random_range, random_range)
        return x + x_random, y + y_random

    def start_execute(self):
        """执行重复次数"""
        start_position, end_position, start_random, stop_random, duration_ = self.parsing_ins_dic()
        # 设置随机坐标
        if start_random:
            start_position = self.random_position(start_position)
        if stop_random:
            end_position = self.random_position(end_position)
        # 执行鼠标拖拽
        re_try = self.ins_dic.get("重复次数")
        for _ in range(re_try):
            self.mouse_drag(start_position, end_position, duration_)
            time.sleep(self.time_sleep)


class SaveForm:
    """保存网页表格"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number
        # 网页控制的部分功能
        self.web_option = WebOption(self.out_mes)

    def parsing_ins_dic(self):
        """解析指令字典"""
        element_value = dict(self.ins_dic)["图像路径"]
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "元素类型": parameter_dic_.get("元素类型"),
            "元素值": element_value,
            "工作簿路径": parameter_dic_.get("工作簿"),
            "工作表名称": parameter_dic_.get("工作表"),
            "超时类型": parameter_dic_.get("异常"),
        }

    def start_execute(self):
        """执行重复次数"""
        list_ins_ = self.parsing_ins_dic()
        # 执行网页操作
        global DRIVER
        self.web_option.driver = DRIVER
        self.web_option.single_shot_operation(
            action="保存表格",
            element_value_=list_ins_.get("元素值"),
            element_type_=list_ins_.get("元素类型"),
            timeout_type_=list_ins_.get("超时类型"),
        )
        self.out_mes.out_mes("已执行保存网页表格", self.is_test)


class ToggleFrame:
    """切换frame"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number
        # 网页控制的部分功能
        self.web_option = WebOption(self.out_mes)

    def parsing_ins_dic(self):
        """解析指令字典"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        instruction_type = parameter_dic_.get("指令类型")
        list_dic = {
            "切换类型": instruction_type,
            "frame类型": parameter_dic_.get("frame类型") if instruction_type == "切换到指定frame" else None,
            "frame值": parameter_dic_.get("frame值") if instruction_type == "切换到指定frame" else None,
        }
        return list_dic

    def start_execute(self):
        """执行重复次数"""
        list_ins_ = self.parsing_ins_dic()
        # 执行网页操作
        self.web_option.switch_to_frame(
            iframe_type=list_ins_.get("frame类型"),
            iframe_value=list_ins_.get("frame值"),
            switch_type=list_ins_.get("切换类型"),
        )
        self.out_mes.out_mes("已执行切换frame", self.is_test)


class SwitchWindow:
    """切换网页窗口"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number
        # 网页控制的部分功能
        self.web_option = WebOption(self.out_mes)

    def parsing_ins_dic(self):
        """解析指令字典"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        list_dic = {
            "切换类型": parameter_dic_.get("窗口类型"),
            "窗口值": parameter_dic_.get("窗口"),
        }
        return list_dic

    def start_execute(self):
        """执行重复次数"""
        list_ins_ = self.parsing_ins_dic()
        # 执行网页操作
        self.web_option.switch_to_window(
            window_type=list_ins_.get("切换类型"), window_value=list_ins_.get("窗口值")
        )
        self.out_mes.out_mes("已执行切换窗口", self.is_test)


class DragWebElements:
    """拖拽网页元素"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False
        self.cycle_number = cycle_number
        # 网页控制的部分功能
        self.web_option = WebOption(self.out_mes)

    def parsing_ins_dic(self):
        """解析指令字典"""
        element_value = dict(self.ins_dic)["图像路径"]
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "元素类型": parameter_dic_.get("元素类型"),
            "元素值": element_value,
            "x": int(parameter_dic_.get('距离X')),
            "y": int(parameter_dic_.get('距离Y')),
            "超时类型": parameter_dic_.get("异常"),
        }

    def start_execute(self):
        """执行重复次数"""
        list_ins_ = self.parsing_ins_dic()
        # 执行网页操作
        global DRIVER
        self.web_option.driver = DRIVER
        self.web_option.distance_x = int(dict(list_ins_)["x"])
        self.web_option.distance_y = int(dict(list_ins_)["y"])
        self.web_option.single_shot_operation(
            action="拖动元素",
            element_value_=list_ins_.get("元素值"),
            element_type_=list_ins_.get("元素类型"),
            timeout_type_=list_ins_.get("超时类型"),
        )
        self.out_mes.out_mes("已执行拖拽网页元素", self.is_test)


class FullScreenCapture:
    """全屏截图"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number

    def parsing_ins_dic(self):
        """解析指令字典"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        save_path = self.ins_dic.get("图像路径")
        return parameter_dic_, save_path

    @staticmethod
    def take_screenshot(screenshot_type, region=None):
        if screenshot_type == "区域截图" and region:
            screenshot = pyautogui.screenshot(region=eval(region))
        else:
            screenshot = pyautogui.screenshot()
        return screenshot

    @staticmethod
    def save_screenshot(screenshot, save_path):
        screenshot.save(save_path)
        return f"已保存截图到{save_path}"

    @staticmethod
    def copy_screenshot_to_clipboard(screenshot):
        # 截图复制到剪切板
        output = io.BytesIO()
        screenshot.convert("RGB").save(output, "BMP")
        data = output.getvalue()[14:]
        output.close()
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(win32con.CF_DIB, data)
        win32clipboard.CloseClipboard()
        return "已复制截图到剪切板"

    def start_execute(self):
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        save_path = self.ins_dic.get("图像路径")
        screenshot_type = parameter_dic_.get("截图类型")
        screenshot = self.take_screenshot(screenshot_type, parameter_dic_.get("区域"))
        if parameter_dic_.get("截图后") == "保存到路径":
            message = self.save_screenshot(screenshot, save_path)
        else:
            message = self.copy_screenshot_to_clipboard(screenshot)
        self.out_mes.out_mes(message, self.is_test)


class SendWeChat:
    """发送微信消息"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = float(get_setting_data_from_ini("Config", "暂停时间"))
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        # 是否是测试
        self.is_test = False
        self.cycle_number = cycle_number

    def parsing_ins_dic(self):
        """解析指令字典"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "联系人": parameter_dic_.get("联系人"),
            "消息内容": parameter_dic_.get("消息内容"),
        }

    @staticmethod
    def get_pid(name):
        """
        作用：根据进程名获取进程pid
        返回：返回匹配第一个进程的pid
        """
        pids = psutil.process_iter()
        for pid in pids:
            if pid.name() == name:
                return pid.pid

    def send_message_to_wechat(self, contact_person, message, repeat_times=1):
        """向微信好友发送消息
        :param contact_person: 联系人
        :param message: 消息内容
        :param repeat_times: 重复次数"""

        def get_correct_message():
            """获取正确的消息内容"""
            if message == "从剪切板粘贴":
                return pyperclip.paste()
            elif message == "当前日期时间":
                return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
            else:
                return sub_variable(message)

        def output_info(judge, message_=None, failure_info=None):
            """向主窗口或na输出提示信息
            :param failure_info:失败信息
            :param judge: （成功、失败）
            :param message_: 消息内容，可选"""
            output_message = None
            if judge == "成功":
                output_message = (
                    f"微信已发送消息：{message_}" if message_ else f"已发送消息"
                )
            elif judge == "失败":
                output_message = f"{failure_info}"
            self.out_mes.out_mes(output_message, self.is_test)

        hwnd = self.get_pid("WeChat.exe")  # 获取微信的后台进程，检查微信是否在运行
        new_message = get_correct_message()
        try:
            if hwnd is not None:
                pyautogui.hotkey("ctrl", "alt", "w")  # 打开微信窗口
                app = Application(backend="uia").connect(process=hwnd)
                wechat_window = app.window(class_name="WeChatMainWndForPC")

                # 找到指定联系人并点击
                # 定位到主窗口
                wx_win = app.window(class_name='WeChatMainWndForPC')
                wx_chat_win = wx_win.child_window(title=contact_person, control_type="ListItem")
                # 聚焦到所需的对话框
                wx_chat_win.click_input()

                for i in range(repeat_times):  # 重复次数
                    pyperclip.copy(new_message)  # 将消息内容复制到剪切板
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.press("enter")  # 模拟按下键盘enter键，发送消息
                    time.sleep(self.time_sleep)

                wechat_window.minimize()  # 最小化窗口
                output_info("成功", new_message)  # 向主窗口输出提示信息
            else:
                output_info(
                    "失败", new_message, "未找到微信窗口，发送失败。"
                )  # 向主窗口输出提示信息
        except Exception as e:
            print(e)
            output_info(
                "失败", new_message, f"发送失败，错误信息：{str(e)}"
            )  # 向主窗口输出提示信息

    def start_execute(self):
        """执行重复次数"""
        list_ins_ = self.parsing_ins_dic()
        re_try = self.ins_dic.get("重复次数")
        # 执行滚轮滑动
        if re_try == 1:
            self.send_message_to_wechat(
                list_ins_.get("联系人"), list_ins_.get("消息内容")
            )
        elif re_try > 1:
            self.send_message_to_wechat(
                list_ins_.get("联系人"), list_ins_.get("消息内容"), re_try
            )


class VerificationCode:

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 主窗口
        self.out_mes = outputmessage
        # 指令字典
        self.ins_dic = ins_dic
        # 网页控制的部分功能
        self.web_option = WebOption(self.out_mes)
        # 是否是测试
        self.is_test = False
        self.cycle_number = cycle_number
        # 云码平台
        self._custom_url = "http://api.jfbym.com/api/YmServer/customApi"
        self._token = get_setting_data_from_ini("三方接口", "云码Token")
        self._headers = {"Content-Type": "application/json"}

    def parsing_ins_dic(self):
        """解析指令字典"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "区域": eval(parameter_dic_.get("区域")),
            "变量": parameter_dic_.get("变量"),
            "验证码类型": parameter_dic_.get("验证码类型"),
        }

    def common_verify(self, image, verify_type="通用数英1-4位"):
        verification_code_types = {
            "通用数英1-4位": 10110,
            "通用数英5-8位": 10111,
            "通用数英9~11位": 10112,
            "通用数英12位及以上": 10113,
            "通用数英1~6位plus": 10103,
            "定制-数英5位~qcs": 9001,
            "定制-纯数字4位": 193,
            "通用中文字符1~2位": 10114,
            "通用中文字符3~5位": 10115,
            "通用中文字符6~8位": 10116,
            "通用中文字符9位及以上": 10117,
            "定制-XX西游苦行中文字符": 10107,
            "通用数字计算题": 50100,
            "通用中文计算题": 50101,
        }
        # 将verify_type转换为对应的数字
        verify_type_int = verification_code_types.get(verify_type)
        payload = {
            "image": base64.b64encode(image).decode(),
            "token": self._token,
            "type": str(verify_type_int),
        }
        resp = requests.post(
            self._custom_url, headers=self._headers, data=json.dumps(payload)
        )
        code = int(resp.json()["code"])
        try:
            result = resp.json()["data"]["data"]
        except Exception as e:
            print(e)
            result = ""
        return code, result

    def ver_input(self, region, verify_type_, variable_name):
        """截图区域，识别验证码，输入验证码"""
        im = pyautogui.screenshot(region=region)
        im_path = os.path.join(os.getcwd(), "ver.png")
        im.save(os.path.join(os.getcwd(), "ver.png"))
        # 使用base64编码
        im_base64 = open(im_path, "rb").read()
        code_, res = self.common_verify(image=im_base64, verify_type=verify_type_)
        if code_ == 10000:
            self.out_mes.out_mes(f"识别出的验证码为：{res}", self.is_test)
        elif code_ == 10001:
            self.out_mes.out_mes(
                "识别验证码失败，账户错误！请设置Token。", self.is_test
            )
        # 释放资源
        os.remove(im_path)
        if not self.is_test:  # 非测试模式下
            # 执行变量写入
            set_variable_value(variable_name, res)
            self.out_mes.out_mes(f"已将识别结果写入变量：{variable_name}", self.is_test)

    def start_execute(self):
        """执行重复次数"""
        list_dic = self.parsing_ins_dic()
        # 执行验证码输入
        self.ver_input(
            list_dic.get("区域"), list_dic.get("验证码类型"), list_dic.get("变量")
        )


class PlayVoice:
    """播放声音"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        self.time_sleep = float(get_setting_data_from_ini("Config", "暂停时间"))
        self.out_mes = outputmessage  # 用于输出信息
        self.ins_dic = ins_dic  # 指令字典
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def play_voice(self):
        """播放声音"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        type_ = parameter_dic_.get("类型")
        if type_ == "系统提示音":
            voice_type = parameter_dic_.get("提示类型")
            self.out_mes.out_mes(f"播放系统提示音：{voice_type}", self.is_test)
            self.system_prompt_tone(sound_type=voice_type)
        elif type_ == "音频信号":
            frequency = int(parameter_dic_.get("频率"))
            self.out_mes.out_mes(f"播放音频信号：{frequency}Hz", self.is_test)
            self.sound_signal(
                frequency=frequency,
                duration=int(parameter_dic_.get("持续")),
                times=int(parameter_dic_.get("次数")),
                interval=int(parameter_dic_.get("间隔")),
            )
        elif type_ == "播放语音":
            text = sub_variable(parameter_dic_.get("内容"))
            self.out_mes.out_mes(f"播放语音：{text}", self.is_test)
            self.play_audio(
                info=text,
                rate=int(parameter_dic_.get("语速")),
            )

    def start_execute(self):
        """开始执行鼠标点击事件"""
        reTry = self.ins_dic.get("重复次数")
        for _ in range(reTry):
            self.play_voice()
            time.sleep(self.time_sleep)

    @staticmethod
    def system_prompt_tone(sound_type) -> None:
        """系统提示音
        :param sound_type: 提示音类型(1:警告, 2:错误, 3:询问, 4:信息, 5:系统启动, 6:系统关闭)"""
        if sound_type == "系统警告":
            winsound.PlaySound("SystemAsterisk", winsound.SND_ALIAS)
        elif sound_type == "系统错误":
            winsound.PlaySound("SystemExclamation", winsound.SND_ALIAS)
        elif sound_type == "系统询问":
            winsound.PlaySound("SystemQuestion", winsound.SND_ALIAS)
        elif sound_type == "系统信息":
            winsound.PlaySound("SystemHand", winsound.SND_ALIAS)
        elif sound_type == "系统启动":
            winsound.PlaySound("SystemStart", winsound.SND_ALIAS)
        elif sound_type == "系统关闭":
            winsound.PlaySound("SystemExit", winsound.SND_ALIAS)

    @staticmethod
    def sound_signal(
            frequency: int, duration: int, times: int = 1, interval: int = 0
    ) -> None:
        """播放音频信号
        :param frequency: 频率(37~32767)
        :param duration: 持续时间(毫秒)
        :param times: 次数
        :param interval: 间隔时间(毫秒)"""
        print(frequency, duration, times, interval)
        try:
            for _ in range(times):
                winsound.Beep(frequency, duration)
                if interval:
                    time.sleep(interval / 1000)
        except RuntimeError:
            print("播放音频信号失败")

    @staticmethod
    def play_audio(info: str, rate: int = 200) -> None:
        """播放TTS提示音"""
        try:
            engine = pyttsx4.init()
            engine.setProperty("rate", rate)  # 设置语速
            engine.say(info)
            engine.runAndWait()
        except Exception as e:
            print(e)


class WaitWindow:
    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        self.root = tk.Tk()
        # 设置标签1
        self.font = ("微软雅黑", 12)  # 设置字体为微软雅黑
        self.label = tk.Label(
            self.root, text="", font=self.font, fg="blue"
        )  # 设置字体和颜色
        self.label.pack(pady=1)
        # 设置标签2
        self.font = ("微软雅黑", 50)  # 设置字体为微软雅黑，字体大小为9
        self.label_2 = tk.Label(
            self.root, text="", font=self.font, fg="red"
        )  # 设置字体和颜色
        self.label_2.pack(pady=1)
        # 设置按钮
        style = ttk.Style()
        style.configure(
            "RoundedButton.TButton",
            font=("Arial", 12, "bold"),
            borderwidth=0,
            relief=tk.RAISED,
        )
        self.button = ttk.Button(
            self.root,
            text="结束等待",
            style="RoundedButton.TButton",
            command=self.stop_win,
        )
        self.button.pack(pady=1)
        # 其他设置
        self.out_mes = outputmessage  # 用于输出信息
        self.ins_dic = ins_dic
        self.cycle_number = cycle_number
        self.count = int(eval(self.ins_dic.get("参数1（键鼠指令）")).get("秒数"))
        self.update_label()
        self.is_test = False
        self.root.protocol("WM_DELETE_WINDOW", self.stop_win)  # 点击关闭按钮时执行
        # 窗口设置
        self.root.geometry("300x200")
        self.root.resizable(False, False)  # 设置禁止调整窗口大小
        self.root.attributes("-toolwindow", 2)
        self.root.attributes("-topmost", True)  # 置顶窗口
        self.root.attributes("-alpha", 0.9)  # 设置透明度
        self.set_screen_to_center()

    def set_screen_to_center(self):
        # 移动窗口到屏幕中央
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = int((screen_width - 300) / 2)
        y = int((screen_height - 200) / 2)
        self.root.geometry("+{}+{}".format(x, y))

    def parsing_ins_dic(self):
        """解析指令字典"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        list_dic = {
            "窗口标题": parameter_dic_.get("标题"),
            "提示信息": parameter_dic_.get("内容"),
            "等待时间": int(parameter_dic_.get("秒数")),
        }
        return list_dic

    def update_label(self):
        if self.count < 1:
            self.root.destroy()
            self.out_mes.out_mes("已结束等待窗口", self.is_test)
            return

        self.label_2.config(text="{}".format(self.count))
        self.count -= 1
        self.root.after(1000, self.update_label)

    def stop_win(self):
        self.root.destroy()
        self.out_mes.out_mes("已结束等待窗口", self.is_test)

    def start_execute(self):
        """开始执行窗口等待"""
        dict_ = self.parsing_ins_dic()  # 解析指令字典
        self.out_mes.out_mes("正在运行等待窗口...", self.is_test)
        self.root.title(dict_.get("窗口标题"))
        self.label.config(text=dict_.get("提示信息"))
        self.root.mainloop()


class DialogWindow:
    """提示框功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic = ins_dic  # 指令字典
        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "标题": parameter_dic_.get("标题"),
            "内容": parameter_dic_.get("内容"),
            "图标": parameter_dic_.get("图标"),
        }

    def start_execute(self):
        """开始执行鼠标点击事件"""
        # 解析指令字典
        ins_dic = self.parsing_ins_dic()
        self.alert_dialog_box(
            ins_dic.get("内容"), ins_dic.get("标题"), ins_dic.get("图标")
        )

    def alert_dialog_box(self, text, title, icon_):
        """弹出对话框
        :param text: 弹窗内容
        :param title: 弹窗标题
        :param icon_: 弹窗图标"""
        self.out_mes.out_mes("已执行弹窗", self.is_test)
        icon_dic = {
            "STOP": pymsgbox.STOP,
            "WARNING": pymsgbox.WARNING,
            "INFO": pymsgbox.INFO,
            "QUESTION": pymsgbox.QUESTION,
        }
        pymsgbox.alert(text=text, title=title, icon=icon_dic.get(icon_))


class BranchJump:
    """跳转分支的功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic = ins_dic  # 指令字典

        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def start_execute(self):
        """开始执行鼠标点击事件"""
        raise IndexError


class TerminationProcess:
    """终止流程的功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic = ins_dic  # 指令字典

        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def start_execute(self):
        """开始执行鼠标点击事件"""
        raise IndexError


class WindowControl:
    """窗口控制"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic = ins_dic  # 指令字典

        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "窗口标题": parameter_dic_.get("标题包含"),
            "操作类型": parameter_dic_.get("操作"),
            "是否报错": eval(parameter_dic_.get("报错")),
        }

    def start_execute(self):
        """开始执行鼠标点击事件"""
        list_dic = self.parsing_ins_dic()
        self.show_normal_window_with_specified_title(
            list_dic.get("窗口标题"), list_dic.get("操作类型"), list_dic.get("是否报错")
        )

    def show_normal_window_with_specified_title(
            self, title, judge="最大化", is_error=True
    ):
        """将指定标题的窗口置顶
        :param is_error: 是否报错
        :param title: 指定标题
        :param judge: 判断（最大化、最小化、显示窗口、关闭）"""

        def get_all_window_title():
            """获取所有窗口句柄和窗口标题"""
            hwnd_title_ = dict()

            def get_all_hwnd(hwnd, mouse):
                if (
                        win32gui.IsWindow(hwnd)
                        and win32gui.IsWindowEnabled(hwnd)
                        and win32gui.IsWindowVisible(hwnd)
                ):
                    hwnd_title_.update({hwnd: win32gui.GetWindowText(hwnd)})

            win32gui.EnumWindows(get_all_hwnd, 0)
            return hwnd_title_

        hwnd_title = get_all_window_title()
        for h, t in hwnd_title.items():
            if title in t:
                print(t)
                print(judge)
                if judge == "最大化":
                    win32gui.ShowWindow(h, win32con.SW_SHOWMAXIMIZED)  # 最大化显示窗口
                elif judge == "最小化":
                    win32gui.ShowWindow(h, win32con.SW_SHOWMINIMIZED)
                elif judge == "显示窗口":
                    win32gui.ShowWindow(h, win32con.SW_SHOWNORMAL)  # 显示窗口
                elif judge == "关闭窗口":
                    win32gui.PostMessage(h, win32con.WM_CLOSE, 0, 0)
                self.out_mes.out_mes(
                    f"已{judge}指定标题包含“{title}”的窗口", self.is_test
                )
                break
        else:
            self.out_mes.out_mes(f"没有找到标题包含“{title}”的窗口！", self.is_test)
            if is_error:
                raise ValueError(f"没有找到标题包含“{title}”的窗口！")


class KeyWait:
    """按键等待的功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic = ins_dic  # 指令字典

        self.is_test = False  # 是否测试
        self.cycle_number = cycle_number  # 循环次数

    def start_execute(self):
        """开始执行鼠标点击事件"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        key = parameter_dic_.get("按键")
        type_ = parameter_dic_.get("等待类型")
        self.out_mes.out_mes(f"等待按键{key}按下中...", self.is_test)
        if type_ == "按键等待":
            keyboard.wait(key.lower())
            self.out_mes.out_mes(f"按键{key}已被按下", self.is_test)
        elif type_ == "跳转分支":
            keyboard.wait(key.lower())
            self.out_mes.out_mes(f"按键{key}已被按下！跳转分支。", self.is_test)
            raise ValueError(f"按键{key}已被按下！")


class GetTimeValue:
    """获取时间变量指令"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "时间格式": parameter_dic_.get("时间格式"),
            "变量名称": parameter_dic_.get("变量"),
        }

    def start_execute(self):
        """开始执行鼠标点击事件"""
        list_dic = self.parsing_ins_dic()  # 参数字典
        now_time_str = str(self.get_now_time(list_dic.get("时间格式")))
        if not self.is_test:
            set_variable_value(list_dic.get("变量名称"), now_time_str)
            self.out_mes.out_mes(f'获取到的值为：{now_time_str}', self.is_test)
            self.out_mes.out_mes(
                f'已获取当前时间并赋值给变量：{list_dic.get("变量名称")}', self.is_test
            )
        else:
            self.out_mes.out_mes(f"已获取当前时间：{now_time_str}", self.is_test)

    @staticmethod
    def get_now_time(format_="年-月-日 小时:分钟:秒"):
        """获取当前时间
        :param format_: 时间格式
        :return: 当前时间字符串"""
        allowed_formats = {
            "年-月-日 小时:分钟:秒": "%Y-%m-%d %H:%M:%S",
            "年/月/日 小时:分钟:秒": "%Y/%m/%d %H:%M:%S",
            "月/日/年 小时:分钟:秒": "%m/%d/%Y %H:%M:%S",
            "日-月-年 小时:分钟:秒": "%d-%m-%Y %H:%M:%S",
            "年-月-日": "%Y-%m-%d",
            "月/日/年": "%m/%d/%Y",
            "日-月-年": "%d-%m-%Y",
            "年-月": "%Y-%m",
            "月/年": "%m/%Y",
            "年": "%Y",
            "时间戳": "%s",  # 时间戳格式
        }

        if format_ not in allowed_formats:
            raise ValueError(
                "Invalid format_. " "Please use one of the allowed formats."
            )
        if format_ == "时间戳":
            return int(time.time())
        else:
            return time.strftime(allowed_formats[format_], time.localtime())


class GetExcelCellValue:
    """从excel表格中获取单元格的值"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        excel_path = get_available_path(self.ins_dic.get("图像路径"), self.out_mes, self.is_test)
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "工作簿路径": excel_path,
            "工作表名称": parameter_dic_.get("工作表"),
            "单元格位置": parameter_dic_.get("单元格"),
            "变量名称": parameter_dic_.get("变量"),
            "行号递增": eval(parameter_dic_.get("递增")),
        }

    def start_execute(self):
        """开始执行鼠标点击事件"""
        list_dic = self.parsing_ins_dic()
        cell_position = list_dic.get("单元格位置")

        if list_dic.get("行号递增"):  # 递增行号
            cell_position = line_number_increment(cell_position, self.cycle_number - 1)

        cell_value = self.get_value_from_excel(
            list_dic.get("工作簿路径"), list_dic.get("工作表名称"), cell_position
        )
        if cell_value is None:
            self.out_mes.out_mes("异常，未获取到单元格的值。", self.is_test)
        else:
            self.send_out_message(cell_value, list_dic)

    def send_out_message(self, cell_value, list_dic):
        if not self.is_test:
            set_variable_value(list_dic.get("变量名称"), cell_value)
            self.out_mes.out_mes(
                f'已获取单元格的值并赋值给变量：{list_dic.get("变量名称")}',
                self.is_test,
            )
        else:
            self.out_mes.out_mes(f"已获取单元格的值：{cell_value}", self.is_test)

    @staticmethod
    def get_value_from_excel(file_path, sheet_name, cell="A1"):
        """从excel中获取数据"""
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb[sheet_name]
            value = sheet[cell].value
            return value
        except Exception as e:
            print(e)
            return None


class GetDialogValue:
    """从对话框中获取值"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "对话框标题": parameter_dic_.get("标题"),
            "变量名称": parameter_dic_.get("变量"),
            "对话框提示信息": parameter_dic_.get("提示"),
        }

    def start_execute(self):
        """开始执行鼠标点击事件"""
        ins_dic = self.parsing_ins_dic()  # 解析指令字典
        text = self.gets_text_from_dialog(ins_dic)
        set_variable_value(ins_dic.get("变量名称"), text)  # 执行变量写入
        self.out_mes.out_mes(
            f'已获取对话框的值并赋值给变量：{ins_dic.get("变量名称")}', self.is_test
        )

    @staticmethod
    def gets_text_from_dialog(ins_dic):
        return pymsgbox.prompt(ins_dic.get("对话框提示信息"), ins_dic.get("对话框标题"))


class GetClipboard:
    """获取剪切板的值"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return parameter_dic_.get("变量")

    def start_execute(self):
        """开始执行鼠标点击事件"""
        variable_name = self.parsing_ins_dic()
        text = self.get_clipboard_text()
        if text != "":
            self.out_mes.out_mes(f'已获取剪贴板的值：{text}', self.is_test)
            if not self.is_test:
                set_variable_value(variable_name, text)
                self.out_mes.out_mes(
                    f'已将值赋予变量：{variable_name}', self.is_test
                )
        else:
            self.out_mes.out_mes("异常，未获取到剪贴板的值。", self.is_test)

    @staticmethod
    def get_clipboard_text():
        win32clipboard.OpenClipboard()
        try:
            text = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
        except Exception as e:
            print('获取剪贴板内容失败！', e)
            text = ''
        finally:
            win32clipboard.CloseClipboard()
        return text


class ContrastVariables:
    """变量判断的功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "变量1": parameter_dic_.get("变量1"),
            "变量2": parameter_dic_.get("变量2"),
            "比较符": parameter_dic_.get("比较符"),
            "变量类型": parameter_dic_.get("类型1"),
        }

    def start_execute(self):
        """开始执行鼠标点击事件"""
        ins_dic = self.parsing_ins_dic()
        variable_dic = get_variable_info("dict")  # 获取变量字典
        # 获取变量名称
        variable1_name = ins_dic.get("变量1")
        variable2_name = ins_dic.get("变量2")
        variable_symbol = ins_dic.get("比较符")
        # 获取变量值
        variable1 = variable_dic.get(variable1_name)
        variable2 = variable_dic.get(variable2_name)
        # 执行变量判断
        result = self.comparison_variable(
            variable1, variable_symbol, variable2, ins_dic.get("变量类型")
        )
        # 输出信息
        self.out_mes.out_mes(
            f'变量判断"{variable1_name}{variable_symbol}{variable2_name}"结果：{result}',
            self.is_test,
        )
        if result:
            raise ValueError("变量判断结果为真，跳转分支。")

    @staticmethod
    def comparison_variable(variable1, comparison_symbol, variable2, variable_type):
        """比较变量"""

        def try_parse_date(variable):
            """尝试将变量解析为日期时间对象"""
            try:
                return parse(variable)
            except ValueError:
                return None

        variable1_ = variable1
        variable2_ = variable2
        if variable_type == "日期或时间":
            variable1_ = try_parse_date(variable1)
            variable2_ = try_parse_date(variable2)
        elif variable_type == "数字":
            variable1_ = eval(variable1)
            variable2_ = eval(variable2)
        elif variable_type == "字符串":
            variable1_ = str(variable1)
            variable2_ = str(variable2)

        if comparison_symbol == "=":
            return variable1_ == variable2_
        elif comparison_symbol == "≠":
            return variable1_ != variable2_
        elif comparison_symbol == ">":
            return variable1_ > variable2_
        elif comparison_symbol == "<":
            return variable1_ < variable2_
        elif comparison_symbol == "包含":
            return variable2_ in variable1_


class RunPython:
    """运行python脚本"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic_ = eval(self.ins_dic.get("参数1（键鼠指令）"))
        image_path = self.ins_dic.get("图像路径")
        return {
            "返回名称": parameter_dic_.get("返回值"),
            "变量名称": parameter_dic_.get("变量"),
            "代码": image_path,
        }

    @staticmethod
    def sub_variable_2(text: str):
        """将text中的变量替换为变量值"""
        new_text = text
        if ("☾" in text) and ("☽" in text):
            variable_dic = get_variable_info("dict")
            for key, value in variable_dic.items():
                new_text = new_text.replace(f"☾{key}☽", str(f'"{value}"'))
        return new_text

    def start_execute(self):
        """开始执行鼠标点击事件"""
        ins_dic = self.parsing_ins_dic()
        try:
            # 定义全局命名空间字典
            globals_dict = {}
            python_code = self.sub_variable_2(ins_dic.get("代码"))
            # 在执行代码时，将结果保存到全局命名空间中
            try:
                exec(python_code, globals_dict)
            except Exception as e:
                print(e)
                self.out_mes.out_mes(f"运行失败：{e}", self.is_test)
                raise ValueError(f"运行失败")
            # 从全局命名空间中获取结果
            result = globals_dict.get(ins_dic.get("返回名称"), None)
            if result is not None:
                if not self.is_test:  # 不是测试时,将结果赋值给变量
                    set_variable_value(ins_dic.get("变量名称"), result)
            self.out_mes.out_mes(f"已执行Python，返回：{result}", self.is_test)
        except Exception as e:
            print(e)
            self.out_mes.out_mes(f"运行失败：{e}", self.is_test)


class RunCmd:
    """运行cmd指令功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        cmd = self.ins_dic.get('图像路径')
        return cmd

    def start_execute(self):
        """开始执行鼠标点击事件"""
        cmd = self.parsing_ins_dic()
        # 打开cmd，并执行命令
        try:
            os.system(cmd)
            self.out_mes.out_mes(f"已执行cmd命令：{cmd}", self.is_test)
        except Exception as e:
            print(e)
            self.out_mes.out_mes(f"运行失败：{e}", self.is_test)
            raise ValueError(f"运行失败")


class RunExternalFile:
    """运行外部文件"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        return self.ins_dic.get("图像路径")

    def start_execute(self):
        """开始执行鼠标点击事件"""
        file_path = self.parsing_ins_dic()
        self.run_external_file(file_path)

    def run_external_file(self, file_path):
        """运行外部文件"""
        try:
            os.startfile(file_path)
            self.out_mes.out_mes(f"已运行外部文件：{file_path}", self.is_test)
        except Exception as e:
            print(e)
            self.out_mes.out_mes(f"运行失败：{e}", self.is_test)
            raise ValueError(f"打开文件失败")


class InputCellExcel:
    """输入到excel单元格的功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic = eval(self.ins_dic.get("参数1（键鼠指令）"))
        excel_path = get_available_path(
            self.ins_dic.get("图像路径"), self.out_mes, self.is_test
        )
        return {
            "工作簿路径": excel_path,
            "工作表名称": parameter_dic.get("工作表"),
            "单元格位置": parameter_dic.get("单元格"),
            "是否递增": eval(parameter_dic.get("递增")),
            "输入内容": parameter_dic.get("文本"),
        }

    def start_execute(self):
        """开始执行鼠标点击事件"""
        # 解析指令字典
        list_dic = self.parsing_ins_dic()
        excel_path = list_dic.get("工作簿路径")
        sheet_name = list_dic.get("工作表名称")
        cell_position = list_dic.get("单元格位置")

        if list_dic.get("是否递增"):
            cell_position = line_number_increment(cell_position, self.cycle_number - 1)

        # 输入到excel单元格
        self.input_to_excel(
            excel_path,
            sheet_name,
            cell_position,
            sub_variable(list_dic.get("输入内容")),
        )

    def input_to_excel(self, file_path, sheet_name, cell_position, input_content):
        """输入到excel单元格
        :param file_path: 文件路径
        :param sheet_name: 表名
        :param cell_position: 单元格位置
        :param input_content: 输入内容"""
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb[sheet_name]
            sheet[cell_position] = input_content
            wb.save(file_path)
            wb.close()
            self.out_mes.out_mes(
                f'已将"{input_content}"写入单元格"{cell_position}"', self.is_test
            )
        except PermissionError:
            self.out_mes.out_mes("文件正在被占用或正在被打开，无法写入", self.is_test)
            raise ValueError("文件被占用，无法写入")
        except Exception as e:
            print(e)
            self.out_mes.out_mes(f"写入单元格失败：{e}", self.is_test)
            raise ValueError(f"写入单元格失败")


class TextRecognition:
    """文字识别功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数
        self.transparent_window = TransparentWindow()  # 框选窗口

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic = eval(self.ins_dic.get("参数1（键鼠指令）"))
        return {
            "截图区域": parameter_dic.get("区域", ""),
            "变量名称": parameter_dic.get("变量", "")
        }

    def start_execute(self):
        """开始执行事件"""
        list_dic = self.parsing_ins_dic()
        if list_dic["截图区域"] != "" or list_dic["变量名称"] != "":
            ocr_text = self.ocr_pic(list_dic["截图区域"])  # 识别图片中的文字
            # 显示识别结果
            print('ocr_text', ocr_text)
            if ocr_text is not None:  # 如果识别成功
                self.out_mes.out_mes(f"OCR识别结果：{ocr_text}", self.is_test)
                if not self.is_test:  # 如果不是测试
                    set_variable_value(list_dic["变量名称"], ocr_text)
                    self.out_mes.out_mes(
                        f'已将OCR识别结果赋值给变量：{list_dic["变量名称"]}', self.is_test
                    )
            else:
                self.out_mes.out_mes(
                    "OCR识别失败！检查网络或查看OCR信息是否设置正确。", self.is_test
                )
        else:
            raise ValueError("参数错误！")

    @staticmethod
    def ocr_pic(reigon):
        """文字识别
        :param reigon: 识别区域"""

        def get_result_from_text(text):
            """从识别结果中提取文字信息"""
            return "\n".join(i["words"] for i in text.get("words_result", []))

        im = pyautogui.screenshot(region=eval(reigon))
        # 将截图数据存储在内存中
        im_bytes = io.BytesIO()
        im.save(im_bytes, format="PNG")
        im_b = im_bytes.getvalue()
        # 返回百度api识别文字信息
        try:
            client_info = get_ocr_info()  # 获取百度api信息
            client = AipOcr(
                client_info["appId"], client_info["apiKey"], client_info["secretKey"]
            )
            return get_result_from_text(client.basicGeneral(im_b))
        except Exception as e:
            print(f"Error: {e} 网络错误识别失败")
            return None
        finally:  # 释放内存
            del im
            del im_bytes


class GetMousePositon:
    """获取鼠标位置功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        return eval(self.ins_dic.get("参数1（键鼠指令）"))

    def start_execute(self):
        """开始执行鼠标点击事件"""
        para_dic = self.parsing_ins_dic()
        var = para_dic.get("变量")
        # 获取当前鼠标位置
        mouse_position = (pyautogui.position().x, pyautogui.position().y)
        self.out_mes.out_mes(f"当前鼠标位置：{mouse_position}", self.is_test)
        self.out_mes.out_mes(f"已将当前鼠标位置赋值给变量：{var}", self.is_test)
        str_mouse_position = f"{mouse_position[0]}, {mouse_position[1]}"
        # 设置变量池中的值
        set_variable_value(var, str_mouse_position)


class WindowFocusWait:
    """窗口焦点等待"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def parsing_ins_dic(self):
        """从指令字典中解析出指令参数"""
        parameter_dic_ = eval(self.ins_dic.get('参数1（键鼠指令）'))
        return {
            "窗口标题": parameter_dic_.get("标题包含"),
            "检测频率": float(int(parameter_dic_.get("检测频率")) / 1000),
            '等待类型': parameter_dic_.get('等待类型'),
            '等待时间': int(parameter_dic_.get('等待时间'))
        }

    def start_execute(self):
        """开始执行"""
        list_dic = self.parsing_ins_dic()
        self.check_focus(
            list_dic.get("窗口标题"),
            list_dic.get("等待时间"),
            list_dic.get("检测频率"),
            True if list_dic.get("等待类型") == "等待窗口获取焦点" else False
        )

    def check_focus(
            self,
            window_title_: str,
            timeout: int = 10,
            frequency: float = 0.5,
            wait_for_focus: bool = True
    ):
        """检查窗口是否获得焦点
        :param window_title_: 窗口标题
        :param timeout: 超时时间
        :param frequency: 检查频率
        :param wait_for_focus: True表示等待窗口获取焦点，False表示等待窗口失去焦点"""
        start_time = time.time()
        self.out_mes.out_mes("等待窗口获得焦点中......" if wait_for_focus else "等待窗口失去焦点中......", self.is_test)
        while True:
            active_window = gw.getActiveWindow()
            if active_window is not None:
                if wait_for_focus:
                    if window_title_ in active_window.title:
                        print("应用程序已经获得了焦点")
                        self.out_mes.out_mes("应用窗口已经获得了焦点，等待结束", self.is_test)
                        break
                else:
                    if window_title_ not in active_window.title:
                        print("应用程序已经失去焦点")
                        self.out_mes.out_mes("应用窗口已经失去了焦点，等待结束", self.is_test)
                        break
            else:
                print("没有找到活动窗口")
                self.out_mes.out_mes("没有找到活动窗口", self.is_test)

            # 检查超时
            elapsed_time = time.time() - start_time
            if elapsed_time > timeout:
                self.out_mes.out_mes("窗口等待超时", self.is_test)
                raise TimeoutError("窗口超过指定时间未获取到焦点" if wait_for_focus else "窗口超过指定时间未失去焦点")

            time.sleep(frequency)


class ColorJudgment:
    """颜色判断功能"""

    def __init__(self, outputmessage, ins_dic, cycle_number=1):
        # 设置参数
        self.time_sleep: float = 0.5  # 等待时间
        self.out_mes = outputmessage  # 用于输出信息到不同的窗口
        self.ins_dic: dict = ins_dic  # 指令字典

        self.is_test: bool = False  # 是否测试
        self.cycle_number: int = cycle_number  # 循环次数

    def start_execute(self):
        params = eval(self.ins_dic.get('参数1（键鼠指令）'))
        pixel_coords = eval(params.get('像素坐标'))
        target_color = eval(params.get('目标颜色'))
        tolerance = int(params.get('误差范围'))
        # 判断像素颜色是否匹配
        if pyautogui.pixelMatchesColor(
                pixel_coords[0],
                pixel_coords[1],
                target_color,
                tolerance
        ):
            self.out_mes.out_mes(f"像素颜色匹配成功", self.is_test)
            if not self.is_test:
                raise ValueError("像素颜色匹配")
        else:
            self.out_mes.out_mes(f"像素颜色不匹配！坐标：{pixel_coords}", self.is_test)
