# coding: utf-8
# Copyright (c) [2022] [federalsadler@sohu.com]
# [Clicker] is licensed under Mulan PSL v2.
# You can use this software according to the terms and conditions of the Mulan PSL v2.
# You may obtain a copy of Mulan PSL v2 at:
# http://license.coscl.org.cn/MulanPSL2
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
# EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
# MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
# See the Mulan PSL v2 for more details.
from __future__ import print_function

import collections
import json
import os.path
import re
import time

import openpyxl
from PyQt5 import QtCore, QtWidgets, QtGui
from PyQt5.QtCore import *
from PyQt5.QtCore import QUrl, Qt
from PyQt5.QtGui import QDesktopServices, QPixmap, QFont
from PyQt5.QtWidgets import (
    QMainWindow,
    QTableWidgetItem,
    QHeaderView,
    QDialog,
    QInputDialog,
    QMenu,
    QFileDialog,
    QStyle,
    QStatusBar,
    QMessageBox,
    QApplication,
    QAction,
    QSplashScreen,
)
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from system_hotkey import SystemHotkey

from functions import get_str_now_time, system_prompt_tone, show_normal_window_with_specified_title, is_hotkey_valid, \
    show_window
from icon import Icon
from ini控制 import set_window_size, save_window_size, get_setting_data_from_ini, update_settings_in_ini, \
    get_global_shortcut, writes_to_branch_info, del_branch_info, ini_to_excel, excel_to_ini, get_branch_repeat_times, \
    set_branch_repeat_times, set_current_branch, get_current_branch
from main_work import CommandThread
from 分支执行窗口 import BranchWindow
from 功能类 import close_browser
from 导航窗口功能 import Na
from 数据库操作 import *
from 窗体.about import Ui_About
from 窗体.mainwindow import Ui_MainWindow
from 窗体.参数窗口 import Ui_Param
from 自动更新 import Check_Update, UpdateWindow
from 设置窗口 import Setting
from 资源文件夹窗口 import Global_s
from 软件信息 import CURRENT_VERSION, MAIN_WEBSITE, ISSUE_WEBSITE, QQ_GROUP, QQ, APP_NAME, \
    Github_WEBSITE, DONATE_WEBSITE
from 选择窗体 import ShortcutTable

collections.Iterable = collections.abc.Iterable


# todo: 指令可编译为python代码
# todo: 从微信获取变量
# todo: 可暂时禁用指令功能
# todo: win通知指令
# todo: excel指令集
# todo: 调试模式
# todo: 动作录制功能
# todo: 使用将指定标题的窗口正常显示后会出现菜单栏阴影的问题

# 用户需求
# todo: 绑定窗口指令
# todo: 快捷导入指令，拖动文件到窗口导入指令
# todo: 成功和失败改变变量值的功能
# todo: 鼠标随机移动添加区域限制
# todo: 导航窗口、设置窗口打开时，按全局快捷键也会触发运行
# todo: 指令可以选择执行，表格中使用checkbox控制
# todo: 指令可导出为json
# todo: 鼠标拖动可设置速度
# todo: 后台截图点击指令
# done: 命令添加窗口不能缩小
# done: 图像点击位置可设置随机范围
# done: 网页录入的指令没有替换变量的值

# https://blog.csdn.net/qq_41567921/article/details/134813496

# activate clicker

# pyinstaller -D -w -i clicker.ico Clicker.py --hidden-import=pyttsx4.drivers --uac-admin -y
# pyinstaller -D -i clicker.ico Clicker.py --hidden-import=pyttsx4.drivers --uac-admin -y

# 添加指令的步骤：
# 1. 在导航页的页面中添加指令的控件
# 2. 在导航页的页面中添加指令的处理函数
# 3. 在导航页的treeWidget中添加指令的名称
# 4. 在功能类中添加运行功能


def timer(func):
    def func_wrapper(*args, **kwargs):
        from time import time

        time_start = time()
        result = func(*args, **kwargs)
        time_end = time()
        time_spend = time_end - time_start
        print("%s cost time: %.3f s" % (func.__name__, time_spend))
        return result

    return func_wrapper


class Main_window(QMainWindow, Ui_MainWindow):
    """主窗口"""
    sigkeyhot = pyqtSignal(str, name="sigkeyhot")  # 自定义信号,用于快捷键
    clear_signal = pyqtSignal()  # 自定义信号，textEdit清空信息，防止在全局快捷键调用时程序崩溃
    show_branch_signal = pyqtSignal()  # 自定义信号，显示分支选择窗口，防止在全局快捷键调用时程序崩溃

    def __init__(self):
        super().__init__()
        # 初始化窗体
        self.setupUi(self)
        # 窗口和信息
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)  # 实例化状态栏
        self.icon = Icon()  # 实例化图标
        self.check_file_integrity()  # 检查文件完整性
        self.add_recent_to_fileMenu()  # 将最近文件添加到菜单中
        self.branch_win = BranchWindow(self)  # 分支选择窗口
        # 检查更新
        self.update_thread = Check_Update(self)  # 自动更新线程
        self.update_thread.show_update_signal.connect(self.update_Qmessage)
        self.update_thread.show_update_window_signal.connect(self.update_window)
        is_update = eval(get_setting_data_from_ini("Config", "启动检查更新"))
        if is_update:
            self.check_update_software(False)
        # 显示导不同的窗口
        self.pushButton.clicked.connect(
            lambda: self.show_windows("导航")
        )  # 显示导航窗口
        self.pushButton_3.clicked.connect(
            lambda: self.show_windows("全局")
        )  # 显示全局参数窗口
        self.actions_2.triggered.connect(lambda: self.show_windows("设置"))  # 打开设置
        self.actionabout.triggered.connect(
            lambda: self.show_windows("关于")
        )  # 打开关于窗体
        self.actionhelp.triggered.connect(
            lambda: self.show_windows("说明")
        )  # 打开使用说明
        self.actionk.triggered.connect(
            lambda: self.show_windows("快捷键说明")
        )  # 打开快捷键说明
        # 主窗体表格功能
        self.actionx.triggered.connect(
            lambda: self.save_data("自动保存")
        )  # 保存指令数据
        self.actiona.triggered.connect(
            lambda: self.save_data("excel")
        )  # 导出数据，导出按钮
        self.actionf.triggered.connect(
            lambda: self.data_import("资源文件夹路径")
        )  # 导入数据
        self.actionj.triggered.connect(
            lambda: self.check_update_software(True)
        )
        # 主窗体开始按钮
        self.pushButton_5.clicked.connect(lambda: self.global_shortcut_key("开始线程"))
        self.start_time = None
        self.pushButton_4.clicked.connect(
            lambda: self.show_windows("分支选择")
        )  # 结束任务按钮
        self.pushButton_6.clicked.connect(
            lambda: self.global_shortcut_key("终止线程")
        )  # 结束任务按钮
        self.pushButton_7.clicked.connect(
            lambda: self.global_shortcut_key("暂停和恢复线程")
        )  # 暂停和恢复按钮
        self.toolButton_8.clicked.connect(self.exporting_operation_logs)  # 导出日志按钮
        self.load_branch_to_combobox()  # 加载分支列表
        # 创建和删除分支
        self.toolButton_2.clicked.connect(self.create_branch)
        self.toolButton.clicked.connect(self.delete_branch)
        self.comboBox.currentIndexChanged.connect(self.get_data)
        # 右键菜单
        self.tableWidget.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tableWidget.customContextMenuRequested.connect(self.generateMenu)
        # 指令执行线程
        self.command_thread = CommandThread(self, None)
        self.command_thread.send_message.connect(self.send_message)
        self.command_thread.finished_signal.connect(self.thread_finished)
        # 设置全局快捷键,用于执行指令的终止
        self.clear_signal.connect(self.clear_textEdit)
        self.show_branch_signal.connect(lambda: self.show_windows("分支选择"))
        self.sigkeyhot.connect(self.global_shortcut_key)
        self.hk_stop = SystemHotkey()
        # 加载上次的指令表格
        self.get_data()
        self.tableWidget.installEventFilter(self)  # 安装事件过滤器,重新设置表格的快捷键
        # 加载窗体初始值
        self.load_initialization()

    def load_initialization(self):
        """加载窗体初始值"""

        def check_file_integrity():
            """检查文件完整性"""
            # 检查命令集.db文件是否存在
            if not os.path.exists("命令集.db"):
                QMessageBox.critical(self, "错误", "命令集.db文件不存在！\n请重新下载软件！")
                sys.exit(1)
            # 检查ini文件是否存在
            if not os.path.exists("config.ini"):
                QMessageBox.critical(self, "错误", "config.ini文件不存在！\n请重新下载软件！")
                sys.exit(1)

        set_window_size(self)  # 获取上次退出时的窗口大小
        branch_name = get_current_branch()
        self.comboBox.setCurrentIndex(self.comboBox.findText(branch_name) if branch_name else 0)
        # 缩小tableWidget行高
        self.tableWidget.verticalHeader().setDefaultSectionSize(20)
        check_file_integrity()  # 检查文件完整性
        # 显示工具栏
        judge = eval(get_setting_data_from_ini("Config", "显示工具栏"))
        self.toolBar.setVisible(judge)
        self.actiong.setChecked(judge)
        # 注册全局快捷键
        self.register_global_shortcut_keys()
        # 设置状态栏信息
        self.statusBar.showMessage(f"软件版本：{CURRENT_VERSION}准备就绪...", 3000)

    def check_file_integrity(self):
        """检查文件完整性"""
        app_path = os.getcwd()
        # 检查ini文件是否存在
        if not os.path.exists(os.path.join(app_path, "config.ini")):
            QMessageBox.critical(self, '致命错误', 'config.ini文件不存在！请重新下载！')
            sys.exit(1)
        # 检查命令集.db文件是否存在
        if not os.path.exists(os.path.join(app_path, "命令集.db")):
            QMessageBox.critical(self, '致命错误', '命令集.db文件不存在！请重新下载！')
            sys.exit(1)
        # 检查开屏和qss文件夹是否存在
        if not os.path.exists(os.path.join(app_path, 'flat')):
            QMessageBox.critical(self, '致命错误', 'flat文件夹不存在！')
            sys.exit(1)
        # 检查qss文件夹下是否有模型文件
        model_files = os.listdir(os.path.join(app_path, 'flat'))
        if not any(x.endswith('.qss') for x in model_files) or not any(x.endswith('.png') for x in model_files):
            QMessageBox.critical(self, '致命错误', 'flat文件夹下没有文件！')
            sys.exit(1)

    def register_global_shortcut_keys(self):
        """注册全局快捷键"""
        # 从ini文件中获取全局快捷键
        global_shortcut = get_global_shortcut()
        # 检查快捷键是否有效，无效则弹出提示
        try:
            global_shortcuts = {
                "开始运行": "开始线程",
                "结束运行": "终止线程",
                "暂停和恢复": "暂停和恢复线程",
                "分支选择": "弹出分支选择窗口"
            }

            for shortcut_name, action in global_shortcuts.items():
                # 将ctrl替换为control
                global_shortcut[shortcut_name] = [
                    key.replace("ctrl", "control") for key in global_shortcut[shortcut_name]
                ]
                if is_hotkey_valid(self.hk_stop, global_shortcut[shortcut_name]):
                    self.hk_stop.register(
                        global_shortcut[shortcut_name],
                        callback=lambda x, action=action: self.global_shortcut_key(action),
                        overwrite=True
                    )
                else:
                    str_shortcut = "+".join(global_shortcut[shortcut_name])
                    QMessageBox.information(
                        self,
                        "提醒",
                        f"快捷键{str_shortcut}已被占用！“{shortcut_name}”的全局快捷键已失效！"
                        f"\n\n请在设置窗口中重新设置全局快捷键。",
                    )
                # 将主界面的按钮显示为快捷键
                self.pushButton_5.setText(f"开始运行\t{'+'.join(global_shortcut['开始运行'])}".upper())
                self.pushButton_4.setText(f"选择分支运行\t{'+'.join(global_shortcut['分支选择'])}".upper())
                self.pushButton_6.setText(f"结束任务\t{'+'.join(global_shortcut['结束运行'])}".upper())
                self.pushButton_7.setText(f"暂停和恢复\t{'+'.join(global_shortcut['暂停和恢复'])}".upper())
        except Exception as e:
            print(e)
            QMessageBox.critical(self, "错误", "全局快捷键已失效！")

    def unregister_global_shortcut_keys(self):
        """注销全局忷键"""
        global_shortcut = get_global_shortcut()
        try:
            for shortcut_name, action in global_shortcut.items():
                # 将ctrl替换为control
                action = [key.replace("ctrl", "control") for key in action]
                self.hk_stop.unregister(tuple(action))
        except Exception as e:
            print(e)

    def add_recent_to_fileMenu(self):
        """将最近文件添加到菜单中"""
        recently_opened_list = get_recently_opened_file("文件列表")
        current_file_path = get_setting_data_from_ini('Config', "当前文件路径")
        # 将最近打开文件添加到菜单中
        if len(recently_opened_list) != 0:
            for file in recently_opened_list:
                file_action = QAction(text=file, parent=self)
                # 设置信号
                file_action.triggered.connect(
                    lambda checked, file_=file: self.open_recent_file(file_)
                )
                file_action.setCheckable(True)
                # 设置当前文件为选中状态
                if file == current_file_path:
                    file_action.setChecked(True)
                self.menuzv.addAction(file_action)
        # 关闭菜单栏
        self.menuzv.close()

    def open_recent_file(self, file_path):
        """打开最近打开的文件
        :param file_path: 文件路径"""
        recent_file = get_setting_data_from_ini('Config', "当前文件路径")
        if file_path != recent_file:
            if os.path.exists(file_path):
                self.data_import(file_path)
            elif not os.path.exists(file_path):
                # 如果文件不存在，则删除最近打开文件列表中的文件
                remove_recently_opened_file(file_path)
                # 从菜单中删除文件
                for action in self.menuzv.actions():
                    if action.text() == file_path:
                        self.menuzv.removeAction(action)
                QMessageBox.critical(
                    self, "错误", "文件不存在！已经从最近打开文件中删除。"
                )
        else:
            for action in self.menuzv.actions():
                if action.text() == file_path:
                    action.setChecked(True)

    def delete_data(self):
        """删除选中的数据行"""
        # 获取选中值的行号和id
        try:
            row = self.tableWidget.currentRow()
            xx = int(self.tableWidget.item(row, 6).text())
            # 删除数据库中指定id的数据
            cursor, con = sqlitedb()
            branch_name = self.comboBox.currentText()
            cursor.execute(
                "delete from 命令 where ID=? and 隶属分支=?",
                (
                    xx,
                    branch_name,
                ),
            )
            con.commit()
            close_database(cursor, con)
            self.get_data(row)  # 调用get_data()函数，刷新表格
            # 状态栏显示信息
            self.statusBar.showMessage(f"删除指令。", 1000)
        except AttributeError:
            pass

    def copy_data(self):
        """复制指定id的指令数据，插入到对应的id位置"""

        def get_new_order():
            """获取新的指令数据"""
            branch_name = self.comboBox.currentText()
            cursor.execute(
                "SELECT * FROM 命令 WHERE ID=? AND 隶属分支=?",
                (
                    id_,
                    branch_name,
                ),
            )
            list_order = cursor.fetchone()
            new_id_ = int(list_order[0]) + 1  # 获取id
            return (new_id_,) + list_order[1:10] + (branch_name,)

        try:
            row = self.tableWidget.currentRow()
            id_ = int(self.tableWidget.item(row, 6).text())  # 指令ID
            cursor, con = sqlitedb()
            new_list_order = get_new_order()  # 获取新的指令数据
            try:
                cursor.execute(
                    "INSERT INTO 命令 VALUES (?,?,?,?,?,?,?,?,?,?,?)", new_list_order
                )
                con.commit()
            except sqlite3.IntegrityError:
                # 如果下一个id已经存在，则将后面的id全部加1
                max_id_ = 1000000
                cursor.execute("UPDATE 命令 SET ID=ID+? WHERE ID>?", (max_id_, id_))
                cursor.execute(
                    "UPDATE 命令 SET ID=ID-? WHERE ID>?",
                    (max_id_ - 1, max_id_ + int(id_)),
                )
                cursor.execute(
                    "INSERT INTO 命令 VALUES (?,?,?,?,?,?,?,?,?,?,?)", new_list_order
                )
                con.commit()
            close_database(cursor, con)
            self.get_data(row)
            self.statusBar.showMessage(f"复制指令。", 1000)
        except AttributeError:
            pass

    def go_to_branch(self):
        """转到分支"""
        row = self.tableWidget.currentRow()  # 获取当前行行号
        branch_name = self.tableWidget.item(row, 2).text()  # 分支名称
        if branch_name not in ["自动跳过", "提示异常并暂停", "提示异常并停止"]:
            # 跳转到对应的分支表和行
            go_branch_name = branch_name.split("-")[0]
            go_row_num = branch_name.split("-")[1]
            self.comboBox.setCurrentText(go_branch_name)
            self.tableWidget.setCurrentCell(int(go_row_num) - 1, 0)  # 设置焦点
        else:
            self.statusBar.showMessage(f"当前指令无分支。", 1000)

    def modify_parameters(self):
        """修改参数"""
        try:
            # 获取当前行行号列号
            row = self.tableWidget.currentRow()
            id_ = int(self.tableWidget.item(row, 6).text())  # 指令ID
            ins_type = self.tableWidget.item(row, 1).text()  # 指令类型
            # 将导航页的tabWidget设置为对应的页
            navigation = Na(self)  # 实例化导航页窗口
            # 修改数据中的参数
            navigation.pushButton_2.setText("修改指令")
            navigation.modify_id = id_
            navigation.show()
            # 获取参数元组：(图像路径，参数，重复次数，异常处理，备注)
            restore_parameters = (
                self.tableWidget.item(row, 0).text(),
                self.tableWidget.item(row, 4).text(),
                self.tableWidget.item(row, 5).text(),
                self.tableWidget.item(row, 2).text(),
                self.tableWidget.item(row, 3).text(),
            )
            navigation.switch_navigation_page(ins_type, restore_parameters)
        except AttributeError:
            QMessageBox.information(self, "提示", "请先选择一行待修改的数据！")

    def move_ins_to_branch(self, branch_name, target_branch_name):
        """移动指令到分支"""
        try:
            row = self.tableWidget.currentRow()
            id_ = int(self.tableWidget.item(row, 6).text())  # 指令ID
            cursor, con = sqlitedb()
            # 获取数据库中id的最大值
            cursor.execute("SELECT MAX(ID) FROM 命令")
            max_id = cursor.fetchone()[0]
            # 将指令移动到目标分支
            if max_id != id_:
                cursor.execute(
                    "UPDATE 命令 SET 隶属分支=?, ID=? WHERE ID=? AND 隶属分支=?",
                    (
                        target_branch_name,
                        max_id + 1,
                        id_,
                        branch_name,
                    ),
                )
            else:
                cursor.execute(
                    "UPDATE 命令 SET 隶属分支=? WHERE ID=? AND 隶属分支=?",
                    (
                        target_branch_name,
                        id_,
                        branch_name,
                    ),
                )
            con.commit()
            close_database(cursor, con)
            self.get_data()
            # 切换到目标分支
            self.comboBox.setCurrentText(target_branch_name)
            # 选中最后一行
            self.tableWidget.setCurrentCell(self.tableWidget.rowCount() - 1, 1)
            self.statusBar.showMessage(f"已将指令移动到分支：{target_branch_name}。", 3000)
        except AttributeError:
            pass

    def open_params_win(self):
        """打开参数窗口"""
        row = self.tableWidget.currentRow()
        params = self.tableWidget.item(row, 4).text()  # parameters
        if (params is not None) and (params != "") and (params != "None"):
            # 格式化字典
            formatted_dict = json.dumps(
                {
                    k: str(v).capitalize() if isinstance(v, bool) else v
                    for k, v in eval(params).items()
                },
                indent=4,
                ensure_ascii=False,
            )
        else:
            formatted_dict = ""
        # 显示参数窗口
        param_win = Param(self)  # create a new window
        param_win.setModal(True)
        param_win.textEdit.setText(formatted_dict)
        param_win.exec_()

    def generateMenu(self, pos):
        """生成右键菜单"""

        def clear_table():
            """清空表格和数据库"""
            choice = QMessageBox.question(self, "提示", "确认清除所有指令吗？")
            if choice == QMessageBox.Yes:
                clear_all_ins()
                # 在ini中删除分支信息，保留主分支
                for i in range(self.comboBox.count()):
                    if self.comboBox.itemText(i) != MAIN_FLOW:
                        del_branch_info(self.comboBox.itemText(i))
                self.get_data()
                self.load_branch_to_combobox()  # 重新加载分支
            else:
                pass

        def insert_data_before(judge):
            """在目标指令前插入指令
            :param judge: （向前插入、向后插入）"""
            try:
                # 获取当前行行号列号
                row = self.tableWidget.currentRow()
                target_id = int(self.tableWidget.item(row, 6).text())  # 指令ID
                navigation = Na(self)  # 实例化导航页窗口
                navigation.show()
                # 修改数据中的参数
                navigation.pushButton_2.setText(judge)
                navigation.modify_id = target_id
                navigation.modify_row = row
            except AttributeError:
                QMessageBox.information(self, "提示", "请先选择一行待修改的数据！")

        # 表格右键菜单
        row_num = -1
        for i_ in self.tableWidget.selectionModel().selection().indexes():
            row_num = i_.row()
        if row_num != -1:  # 未选中数据不弹出右键菜单
            menu = QMenu()  # 实例化菜单

            run_ins = menu.addAction("运行选中指令")
            run_ins.setIcon(
                self.style().standardIcon(QStyle.SP_MediaPlay)
            )

            run_from_this_ins = menu.addAction("从当前行运行")
            run_from_this_ins.setIcon(
                self.style().standardIcon(QStyle.SP_MediaPlay)
            )

            menu.addSeparator()
            refresh = menu.addAction("刷新")
            refresh.setIcon(
                self.style().standardIcon(QStyle.SP_BrowserReload)
            )  # 设置图标

            modify_params = menu.addAction("查看参数")
            modify_params.setIcon(self.icon.view)  # 设置图标

            up_ins = menu.addAction("上移")
            up_ins.setShortcut("Shift+↑")
            up_ins.setIcon(self.style().standardIcon(QStyle.SP_ArrowUp))  # 设置图标

            down_ins = menu.addAction("下移")
            down_ins.setShortcut("Shift+↓")
            down_ins.setIcon(self.style().standardIcon(QStyle.SP_ArrowDown))  # 设置图标

            menu.addSeparator()
            insert_ins_before = menu.addAction("在前面插入指令")
            insert_ins_before.setIcon(self.icon.move_up)  # 设置图标

            insert_ins_after = menu.addAction("在后面插入指令")
            insert_ins_after.setIcon(self.icon.move_down)  # 设置图标

            menu.addSeparator()
            copy_ins = menu.addAction("复制指令")
            copy_ins.setShortcut("Ctrl+C")
            copy_ins.setIcon(self.icon.copy)  # 设置图标

            modify_ins = menu.addAction("修改指令")
            modify_ins.setShortcut("Ctrl+Y")
            modify_ins.setIcon(self.icon.modify_instruction)  # 设置图标

            move_to_branch_menu = menu.addMenu("移动指令到分支")
            move_to_branch_menu.setIcon(self.icon.move_to_branch)
            # 从self.comboBox中获取分支列表
            branch_list = [self.comboBox.itemText(i) for i in range(self.comboBox.count())]
            for branch in branch_list:
                action = move_to_branch_menu.addAction(branch)  # 右键菜单添加分支
                action.triggered.connect(lambda _, b=branch: self.move_ins_to_branch(self.comboBox.currentText(), b))

            menu.addSeparator()
            go_branch = menu.addAction("转到分支")
            go_branch.setShortcut("Ctrl+G")
            go_branch.setIcon(
                self.style().standardIcon(QStyle.SP_ArrowForward)
            )  # 设置图标

            del_ins = menu.addAction("删除指令")
            del_ins.setShortcut("Delete")
            del_ins.setIcon(
                self.style().standardIcon(QStyle.SP_DialogCancelButton)
            )  # 设置图标

            menu.addSeparator()
            del_branch = menu.addAction("删除当前分支指令")
            del_branch.setIcon(
                self.style().standardIcon(QStyle.SP_DialogDiscardButton)
            )  # 设置图标

            del_all_ins = menu.addAction("删除全部指令")
            del_all_ins.setIcon(self.icon.delete)  # 设置图标

            action = menu.exec_(self.tableWidget.mapToGlobal(pos))
        else:
            return

        # 各项操作
        if action == copy_ins:
            self.copy_data()  # 复制指令
        if action == modify_params:
            self.open_params_win()  # 修改指令参数
        elif action == del_ins:
            self.delete_data()  # 删除指令
        elif action == up_ins:
            self.go_up_down("up")
        elif action == down_ins:
            self.go_up_down("down")
        elif action == modify_ins:
            self.modify_parameters()  # 修改指令
        elif action == insert_ins_before:
            insert_data_before("向前插入")
        elif action == insert_ins_after:
            insert_data_before("向后插入")
        elif action == refresh:
            self.get_data()
            self.statusBar.showMessage(f"刷新指令表格。", 1000)
        elif action == del_all_ins:
            clear_table()
            self.statusBar.showMessage(f"清空指令表格。", 1000)
        elif action == del_branch:
            clear_all_ins(branch_name=self.comboBox.currentText())
            self.get_data()
            self.statusBar.showMessage(f"清空当前分支全部指令。", 1000)
        elif action == go_branch:
            self.go_to_branch()
        elif action == run_ins:
            # 获取选中的行的id
            row = self.tableWidget.currentRow()
            id_ = int(self.tableWidget.item(row, 6).text())
            self.start('单行指令', id_)
        elif action == run_from_this_ins:
            # 获取选中行的行号
            row = self.tableWidget.currentRow()
            self.start('从当前行运行', row)

    def show_windows(self, judge):
        """打开窗体"""
        if judge == "设置":
            setting_win = Setting(self)  # 设置窗体
            setting_win.tabWidget.setCurrentIndex(0)
            setting_win.setModal(True)
            setting_win.exec_()
        elif judge == "全局":
            global_s = Global_s(self)  # 全局设置窗口
            global_s.setModal(True)
            global_s.exec_()
        elif judge == "导航":
            navigation = Na(self)  # 实例化导航页窗口
            navigation.show()
        elif judge == "关于":
            about = About(self)  # 设置关于窗体
            about.setModal(True)
            about.exec_()
        elif judge == "分支选择":  # 分支选择窗口
            if not self.branch_win.isVisible():
                self.branch_win.show()
                # 获取焦点
                self.branch_win.activateWindow()
            else:
                self.branch_win.close()
        elif judge == "说明":
            QDesktopServices.openUrl(QUrl(MAIN_WEBSITE))
        elif judge == "快捷键说明":
            title = ["快捷键", "说明"]
            data = [
                ("Ctrl+Enter", "添加指令"),
                ("Ctrl+C", "复制指令"),
                ("Delete", "删除指令"),
                ("Shift+↑", "上移指令"),
                ("Shift+↓", "下移指令"),
                ("Ctrl+↑", "切换到上个分支"),
                ("Ctrl+↓", "切换到下个分支"),
                ("Ctrl+G", "转到分支"),
                ("Ctrl+Y", "修改指令"),
                ("Ctrl+D", "导入指令"),
                ("Ctrl+S", "保存指令"),
                ("Ctrl+Alt+S", "另存为Excel")
            ]
            shortcut_win = ShortcutTable(self, title, data)  # 快捷键说明窗口
            shortcut_win.setModal(True)
            shortcut_win.exec_()

    def get_data(self, row=None):
        """从数据库获取数据并存入表格
        :param row: 设置焦点行号"""
        try:
            self.tableWidget.clearContents()
            self.tableWidget.setRowCount(0)
            # 获取数据库数据
            cursor, con = sqlitedb()
            branch_name = self.comboBox.currentText()
            cursor.execute(
                "select 图像名称,指令类型,异常处理,备注,参数1,重复次数,ID from 命令 where 隶属分支=?",
                (branch_name,),
            )
            list_order = cursor.fetchall()
            close_database(cursor, con)
            # 在表格中写入数据
            for i_ in range(len(list_order)):
                self.tableWidget.insertRow(i_)
                for j in range(len(list_order[i_])):
                    self.tableWidget.setItem(
                        i_, j, QTableWidgetItem(str(list_order[i_][j]))
                    )
            # 自适应列宽（排除第一列和第四列）
            header = self.tableWidget.horizontalHeader()
            for col in range(header.count()):
                if col != 0 and col != 4:
                    header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
            # 设置焦点
            if row is not None:
                self.tableWidget.setCurrentCell(int(row), 0)
            # 设置重复次数
            self.spinBox.setValue(int(get_branch_repeat_times(branch_name)))

        except sqlite3.OperationalError:
            pass

    def go_up_down(self, judge):
        """向上或向下移动选中的行"""

        def database_exchanges_two_rows(id_1: int, id_2: int) -> None:
            """交换数据库中的两行数据
            :param id_1: 要交换的第一行的id
            :param id_2: 要交换的第二行的id"""
            cursor, con = sqlitedb()
            # 交换两行的id
            cursor.execute(
                "update 命令 set ID=? where ID=?",
                (
                    999999,
                    id_1,
                ),
            )
            cursor.execute(
                "update 命令 set ID=? where ID=?",
                (
                    id_1,
                    id_2,
                ),
            )
            cursor.execute(
                "update 命令 set ID=? where ID=?",
                (
                    id_2,
                    999999,
                ),
            )
            con.commit()
            # 刷新数据库
            close_database(cursor, con)

        try:
            # 获取选中值的行号和id
            row = self.tableWidget.currentRow()
            id_ = int(self.tableWidget.item(row, 6).text())
            if judge == "up":
                if row != 0:
                    # 查询上一行的id
                    id_row_up = int(self.tableWidget.item(row - 1, 6).text())
                    # 交换两行的id
                    database_exchanges_two_rows(id_, id_row_up)
                    self.get_data(row - 1)
                    self.statusBar.showMessage(f"上移指令。", 1000)
            elif judge == "down":
                if row != self.tableWidget.rowCount() - 1:
                    # 查询下一行的id
                    id_row_down = int(self.tableWidget.item(row + 1, 6).text())
                    # 交换两行的id
                    database_exchanges_two_rows(id_, id_row_down)
                    self.get_data(row + 1)
                    self.statusBar.showMessage(f"下移指令。", 1000)
        except AttributeError:
            pass

    def save_data(self, judge: str):
        """保存配置文件到当前文件夹下
        :param judge: 保存的文件类型（excel、自动保存）"""

        def get_save_file_and_folder() -> tuple:
            """获取保存文件名和文件夹路径"""
            # 获取资源文件夹路径作为默认路径，如果存在则使用用户的主目录
            directory_folder_path = extract_resource_folder_path()[0] \
                if extract_resource_folder_path() else os.path.expanduser("~")
            directory_path = os.path.normpath(os.path.join(directory_folder_path, "指令数据.xlsx"))
            # 获取保存文件名和文件夹路径
            file_path, _ = QFileDialog.getSaveFileName(
                parent=self,
                caption="保存文件",
                filter="(*.xlsx)",
                directory=directory_path
            )
            return (os.path.normpath(os.path.split(file_path)[0]),
                    os.path.normpath(os.path.split(file_path)[1])) \
                if file_path else (None, None)

        def get_file_and_folder_from_setting():
            """从设置中获取最近打开的文件路径作为保存路径，用于自动保存"""
            recently_opened = get_setting_data_from_ini('Config', "当前文件路径")
            if recently_opened != "None" and os.path.exists(recently_opened):
                return os.path.split(recently_opened)
            self.statusBar.showMessage("未找到最近导入的文件路径。已自动切换为另存为...", 3000)
            return get_save_file_and_folder()

        def prompt_save_success(save_path_: str):
            """提示保存成功"""
            # 提示保存成功，是否打开文件夹
            if judge != "自动保存" and QMessageBox.question(
                    self, "提示", "指令数据保存成功！是否打开文件夹？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
            ) == QMessageBox.Yes:
                os.startfile(save_path_)
            self.statusBar.showMessage(f"指令数据已保存至{save_path_}。", 3000)

        def adaptive_column_width(sheet_, max_width=50):
            """
            自动设置单元格宽度，并加上最大宽度限制。

            :param sheet_: 工作表对象
            :param max_width: 列宽的最大限制（默认为50）
            """
            for col in range(1, sheet_.max_column + 1):
                max_length = 0
                for cell in sheet_[get_column_letter(col)]:
                    # 计算单元格内容的长度，中文字符的长度为0.7
                    cell_length = (0.7 * len(re.findall(r"([\u4e00-\u9fa5])", str(cell.value)))
                                   + len(str(cell.value)))
                    max_length = max(max_length, cell_length)
                # 设置列宽，但不超过最大宽度
                adjusted_width = min(max_length + 5, max_width)
                sheet_.column_dimensions[get_column_letter(col)].width = adjusted_width

        def set_title_style(sheet_):
            """设置标题样式"""
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            for cell in sheet_[1]:  # 第一行标题
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

        # 判断是否为另存为,如果不是则自动判断文件类型
        try:
            folder_path, file_name = get_save_file_and_folder() \
                if judge != "自动保存" else get_file_and_folder_from_setting()
            # 开始保存数据
            if all([file_name, folder_path]):
                # 使用openpyxl模块创建Excel文件
                wb = openpyxl.Workbook()
                # 获取全局参数表中的分支表名
                branch_table_list = get_branch_info(keys_only=True)
                # 将sheet名设置为分支表名
                headers = [
                    "ID",
                    "图像名称",
                    "指令类型",
                    "参数信息",
                    "参数-2",
                    "参数-3",
                    "参数-4",
                    "重复次数",
                    "异常处理",
                    "备注",
                    "隶属分支"
                ]
                for branch_name in branch_table_list:
                    sheet = wb.create_sheet(branch_name)
                    sheet.append(headers)
                    set_title_style(sheet)  # 设置标题样式
                    # 写入数据
                    for ins in extracted_ins_from_database(branch_name):
                        sheet.append(ins)
                    adaptive_column_width(sheet)

                wb.remove(wb["Sheet"])  # 删除默认的sheet
                ini_to_excel(wb)  # 将ini文件中的数据写入到Excel文件中
                adaptive_column_width(wb['设置'])
                # 保存Excel文件
                save_path = os.path.normpath(os.path.join(folder_path, file_name))
                wb.save(save_path)
                prompt_save_success(save_path)  # 提示保存成功
        except PermissionError:
            QMessageBox.critical(self, "错误", "保存失败，文件被占用！")

    def closeEvent(self, event):
        """关闭窗口事件"""
        # 是否隐藏工具栏
        update_settings_in_ini('Config', 显示工具栏=str(self.actiong.isChecked()))
        # 终止线程
        if self.command_thread.isRunning():
            self.command_thread.terminate()
        # 是否退出清空数据库
        if eval(get_setting_data_from_ini("Config", "退出提醒清空指令")):
            choice = QMessageBox.question(
                self, "提示", "确定退出并清空所有指令？\n将自动保存当前指令数据。"
            )
            if choice == QMessageBox.Yes:
                # 退出终止后台进程并清空数据库
                self.save_data("自动保存")
                event.accept()
                clear_all_ins()
            else:
                event.ignore()
        self.branch_win.close()  # 关闭选择窗口
        # 保存当前分支
        set_current_branch(self.comboBox.currentText())
        # 窗口大小
        save_window_size(self.width(), self.height(), self.windowTitle())

    def data_import(self, file_path):
        """导入数据功能"""

        def data_import_from_excel(target_path_):
            # 读取数据
            wb = openpyxl.load_workbook(target_path_)
            sheets = wb.worksheets  # 获取所有的sheet
            excel_to_ini(wb)  # 写入ini设置
            cursor_, con_ = sqlitedb()
            for sheet in sheets:  # 遍历所有的sheet，写入分支指令
                if sheet.title != "设置":
                    # writes_to_branch_info(sheet.title, '')  # 添加分支表名
                    max_row = sheet.max_row
                    max_column = sheet.max_column
                    # 向数据库中写入数据
                    try:
                        for row in range(2, max_row + 1):
                            # 获取第一列数据
                            instructions = []
                            for column in range(1, max_column + 1):
                                # 获取单元格数据
                                data = sheet.cell(row, column).value
                                instructions.append(data)
                            cursor_.execute(
                                "INSERT INTO 命令(ID,图像名称,指令类型,参数1,参数2,参数3,参数4,"
                                "重复次数,异常处理,备注,隶属分支) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                                instructions[0:11],
                            )
                            con_.commit()
                    except Exception as e:
                        # 捕获并处理异常
                        QMessageBox.warning(self, f"导入失败", f"ID重复或格式错误！{e}")
            wb.close()
            close_database(cursor_, con_)
            self.load_branch_to_combobox()  # 重新加载分支列表
            if file_path == "资源文件夹路径":
                QMessageBox.information(self, "提示", "指令数据导入成功！")

        # 获取资源文件夹路径，如果不存在则使用用户的主目录
        if file_path == "资源文件夹路径":
            directory_path = next(
                (item for item in extract_resource_folder_path()), os.path.expanduser("~")
            )
            target_path, _ = QFileDialog.getOpenFileName(
                parent=self,
                caption="请选择指令备份文件",
                directory=directory_path,
                filter="(*.xlsx)"
            )
            if target_path:
                suffix = os.path.splitext(target_path)[1]
            else:
                return
        else:
            target_path = (file_path, "")
            suffix = os.path.splitext(file_path)[1]

        # 如果为.xlsx文件
        if suffix == ".xlsx":
            clear_all_ins(True)  # 清空原有数据，包括分支表
            data_import_from_excel(target_path)
        # 将最近导入的文件路径写入数据库,用于保存时自动设置路径
        update_settings_in_ini(
            "Config",
            当前文件路径=os.path.normpath(target_path)
        )  # 写入当前文件路径
        writes_to_recently_opened_files(
            os.path.normpath(target_path)
        )  # 写入最近打开的文件
        self.statusBar.showMessage(f"指令数据导入成功！已自动设置保存路径。", 1000)
        self.menuzv.clear()  # 清空最近打开文件菜单
        self.add_recent_to_fileMenu()  # 将最近文件添加到菜单中

    def start(self, run_mode='全部指令', info=0):
        """主窗体开始按钮
        :param run_mode: 运行模式（全部指令、单行指令、从当前行运行）
        :param info: 指令ID"""

        def operation_before_execution():
            """执行前的操作"""
            self.clear_signal.emit()  # 清空日志
            self.tabWidget.setCurrentIndex(0)  # 切换到日志页
            if self.checkBox_2.isChecked():  # 如果勾选了执行中隐藏主窗口
                self.hide()

        if self.command_thread.isRunning():  # 如果线程正在运行,则终止
            self.command_thread.terminate()
        operation_before_execution()  # 执行前的操作
        if run_mode == '全部指令':
            self.command_thread.set_run_mode('全部指令', 0)  # 设置运行模式
            self.command_thread.set_branch_name_index(int(self.comboBox.currentIndex()))
        elif run_mode == '单行指令':
            self.command_thread.set_run_mode('单行指令', info)  # 设置运行模式
            self.command_thread.set_branch_name_index(0)
        elif run_mode == '从当前行运行':
            self.command_thread.set_run_mode('从当前行运行', info)  # 设置运行模式
            self.command_thread.set_branch_name_index(int(self.comboBox.currentIndex()))
        # 设置重复次数
        repeat_number = self.spinBox.value() if self.radioButton_2.isChecked() else -1
        self.command_thread.set_repeat_number(repeat_number)  # 设置重复次数
        set_branch_repeat_times(self.comboBox.currentText(), repeat_number)  # 设置分支重复次数
        # 开始运行
        self.command_thread.start()
        # 记录开始时间的时间戳
        self.start_time = time.time()

    def start_from_branch(self, branch_name, repeat_number=1):
        """从分支开始运行"""
        if self.command_thread.isRunning():  # 如果线程正在运行,则终止
            self.command_thread.terminate()
        self.clear_signal.emit()  # 清空日志
        self.tabWidget.setCurrentIndex(0)  # 切换到日志页
        if self.checkBox_2.isChecked():  # 如果勾选了执行中隐藏主窗口
            self.hide()
        # 获取branch_name在self.comboBox中的索引
        branch_index = self.comboBox.findText(branch_name)
        self.command_thread.set_run_mode('全部指令', 0)  # 设置运行模式
        self.command_thread.set_branch_name_index(branch_index)
        self.command_thread.set_repeat_number(repeat_number)  # 设置重复次数
        set_branch_repeat_times(branch_name, repeat_number)  # 记录分支重复次数
        # 设置主窗口显示的重复次数
        if self.comboBox.currentText() == branch_name:
            self.spinBox.setValue(repeat_number)
        self.command_thread.start()
        # 记录开始时间的时间戳
        self.start_time = time.time()

    def clear_textEdit(self):
        """清空日志，主要用于在全局快捷键线程中调用，避免线程阻塞引发的程序闪退"""
        self.textEdit.clear()

    def exporting_operation_logs(self):
        """导出操作日志"""
        # 打开保存文件对话框
        target_path = QFileDialog.getSaveFileName(
            parent=self,
            caption="请选择保存路径",
            directory=os.path.join(os.path.expanduser("~"), "操作日志.txt"),
            filter="(*.txt)",
        )
        # 判断是否选择了文件
        if target_path[0] != "":
            # 获取操作日志
            logs = self.textEdit.toPlainText()
            # 将操作日志写入文件
            with open(target_path[0], "w") as f:
                f.write(f"日志导出时间：{get_str_now_time()}\n")
                f.write(logs)
            QMessageBox.information(self, "提示", "操作日志导出成功！")

    def create_branch(self):
        """创建分支"""
        flag = Qt.WindowCloseButtonHint
        branch_name, ok = QInputDialog.getText(self, "创建分支", "请输入分支名称：", flags=flag)
        if ok:
            message = writes_to_branch_info(branch_name, '')
            self.load_branch_to_combobox(branch_name)
            QMessageBox.information(
                self, "提示",
                "分支创建成功!" if message else "分支已存在!"
            )

    def delete_branch(self):
        text = self.comboBox.currentText()
        if text == MAIN_FLOW:
            QMessageBox.critical(self, "提示", "无法删除主分支！")
        else:
            # 将combox显示的名称切换为主流程
            self.comboBox.setCurrentIndex(0)
            # 删除分支表
            mes = del_branch_info(text)
            if mes:
                del_branch_in_database(text)  # 删除数据库中的分支
                self.load_branch_to_combobox()  # 重新加载分支列表
                QMessageBox.information(self, "提示", "分支已删除！")
            else:
                QMessageBox.critical(self, "提示", "分支删除失败！")

    def load_branch_to_combobox(self, text=None):
        """加载分支
        :param text: 设置combox的文本"""
        self.comboBox.clear()
        self.comboBox.addItems(get_branch_info(True))
        if text is not None:
            self.comboBox.setCurrentText(text)
        # 设置重复次数
        self.spinBox.setValue(int(get_branch_repeat_times(self.comboBox.currentText())))

    def eventFilter(self, obj, event):
        # 重写self.tableWidget的快捷键事件
        if obj == self.tableWidget:
            if event.type() == 6:  # 键盘按下事件
                # 如果按下delete键
                if event.key() == Qt.Key_Delete:
                    self.delete_data()
                # 如果按下ctrl+c键
                if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_C:
                    self.copy_data()
                # 如果按下shift+向上键
                if event.modifiers() == Qt.ShiftModifier and event.key() == Qt.Key_Up:
                    self.go_up_down("up")
                    # 将焦点下移一行,抵消上移的误差
                    self.tableWidget.setCurrentCell(
                        self.tableWidget.currentRow() + 1, 0
                    )
                # 如果按下shift+向下键
                if event.modifiers() == Qt.ShiftModifier and event.key() == Qt.Key_Down:
                    self.go_up_down("down")
                    # 将焦点上移一行,抵消下移的误差
                    self.tableWidget.setCurrentCell(
                        self.tableWidget.currentRow() - 1, 0
                    )
                # 如果按下ctrl+向上键
                if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_Up:
                    # 将焦点下移一行,抵消上移的误差
                    self.tableWidget.setCurrentCell(
                        self.tableWidget.currentRow() + 1, 0
                    )
                    # 如果分支不为self.comboBox的第一个则,切换上一个分支
                    if self.comboBox.currentIndex() != 0:
                        self.comboBox.setCurrentIndex(self.comboBox.currentIndex() - 1)
                # 如果按下ctrl+向下键
                if (
                        event.modifiers() == Qt.ControlModifier
                        and event.key() == Qt.Key_Down
                ):
                    # 将焦点上移一行,抵消下移的误差
                    self.tableWidget.setCurrentCell(
                        self.tableWidget.currentRow() - 1, 0
                    )
                    # 如果分支不为self.comboBox的最后一个则,切换下一个分支
                    if self.comboBox.currentIndex() != self.comboBox.count() - 1:
                        self.comboBox.setCurrentIndex(self.comboBox.currentIndex() + 1)
                # 如果按下ctrl+g键
                if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_G:
                    self.go_to_branch()  # 转到分支
                # 如果按下ctrl+x键
                if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_Y:
                    self.modify_parameters()  # 修改指令
        return super().eventFilter(obj, event)

        # 热键处理函数

    def global_shortcut_key(self, i_str):
        """全局热键处理函数"""
        system_prompt_tone("全局快捷键")  # 发出提示音

        if i_str == "终止线程":
            if self.command_thread.isRunning():
                self.command_thread.terminate()  # 终止线程
                # 获取当前时间
                self.send_message("任务终止！")
                if self.checkBox_2.isChecked():
                    self.show()
                QApplication.processEvents()
                show_normal_window_with_specified_title(self.windowTitle())  # 显示窗口

        elif i_str == "开始线程":
            self.start('全部指令', 0)  # 开始线程
            self.send_message("任务开始！")

        elif i_str == "暂停和恢复线程":
            if self.command_thread.isRunning():
                if self.command_thread.is_paused:
                    self.send_message("任务恢复！")
                    self.command_thread.resume()
                else:
                    self.send_message("任务暂停！")
                    self.command_thread.pause()

        elif i_str == "弹出分支选择窗口":
            self.show_branch_signal.emit()
            # 将焦点切换到分支选择窗口
            self.branch_win.activateWindow()

    def sendkeyevent(self, i_str):
        """发送热键信号,将外部信号，转化成qt信号,用于全局热键"""
        self.sigkeyhot.emit(i_str)

    def send_message(self, message):
        """向日志窗口发送信息"""
        time_message = f"<font color=#ffff00>{get_str_now_time()}</font>"
        if message != "换行":
            self.textEdit.append(f"{time_message}&nbsp;&nbsp;&nbsp;&nbsp;{message}")
        else:
            self.textEdit.append('')

    def thread_finished(self, message):

        def send_elapsed_time():
            """发送耗时"""
            elapsed_time = time.time() - self.start_time
            # 将秒转换为毫秒或者保留两位小数的秒数
            if elapsed_time < 1:
                elapsed_time_ms = round(elapsed_time * 1000)  # 毫秒
                return f"{elapsed_time_ms}毫秒"
            else:
                elapsed_time_sec = round(elapsed_time, 2)  # 秒，保留两位小数
                return f"{elapsed_time_sec}秒"

        self.send_message(f"{message}，耗时{send_elapsed_time()}。")
        if self.checkBox_2.isChecked():  # 显示窗口
            self.show()
            QApplication.processEvents()
        system_prompt_tone("线程结束")  # 发出提示音
        show_normal_window_with_specified_title(self.windowTitle())  # 显示窗口
        close_browser()  # 关闭浏览器驱动

    def check_update_software(self, show_MessageBox=True):
        """检查更新"""
        self.update_thread.set_show_info(show_MessageBox)
        self.update_thread.start()

    def update_Qmessage(self, message, message_type):
        message_box = {
            "警告": QMessageBox.warning,
            "错误": QMessageBox.critical,
            "信息": QMessageBox.information
        }
        message_box[message_type](self, '提示', message)

    def update_window(self, update_info_dic_):
        """显示更新窗口"""
        update_win = UpdateWindow(self, update_info_dic_)
        update_win.setModal(True)
        update_win.exec_()


class About(QDialog, Ui_About):
    """关于窗体"""

    def __init__(self, parent=None):
        super().__init__(parent)
        # 初始化窗体
        self._parent = parent
        self.setupUi(self)
        self.setWindowFlags(
            self.windowFlags() & ~Qt.WindowContextHelpButtonHint
        )  # 隐藏帮助按钮
        set_window_size(self)  # 获取上次退出时的窗口大小
        self.label_2.setText(f"版本：{CURRENT_VERSION}")  # 设置版本号
        self.label_7.setText('<a href="{}"><font color="red">{}</font></a>'.format(QQ_GROUP, QQ))
        # 绑定事件
        self.gitee.clicked.connect(
            lambda: QDesktopServices.openUrl(QUrl(MAIN_WEBSITE))
        )
        self.gitee_2.clicked.connect(
            lambda: QDesktopServices.openUrl(QUrl(Github_WEBSITE))
        )
        self.pushButton.clicked.connect(lambda: self._parent.check_update_software(True))
        self.pushButton_2.clicked.connect(
            lambda: QDesktopServices.openUrl(QUrl(ISSUE_WEBSITE))
        )
        self.pushButton_3.clicked.connect(
            lambda: QDesktopServices.openUrl(QUrl(DONATE_WEBSITE))
        )

    def closeEvent(self, event):
        # 保存窗体大小
        save_window_size(self.width(), self.height(), self.windowTitle())


class Param(QDialog, Ui_Param):
    """参数设置窗口"""

    def __init__(self, parent=None):
        super().__init__(parent)
        # 初始化窗体
        self.setupUi(self)
        self.setWindowFlags(
            self.windowFlags() & ~Qt.WindowContextHelpButtonHint
        )  # 隐藏帮助按钮
        set_window_size(self)  # 获取上次退出时的窗口大小
        self.pushButton.clicked.connect(self.modify_parameters)  # 保存参数

    def closeEvent(self, event):
        # 保存窗体大小
        save_window_size(self.width(), self.height(), self.windowTitle())

    def modify_parameters(self):
        self.parent().modify_parameters()
        self.close()


class QSSLoader:
    """QSS皮肤加载器"""

    def __init__(self):
        pass

    @staticmethod
    def read_qss_file(qss_file_name):
        """从文件中读取qss的静态方法"""
        with open(qss_file_name, "r", encoding="UTF-8") as file:
            return file.read()


if __name__ == "__main__":
    # 自适应高分辨率
    # 强制启用高 DPI 感知模式
    # 需要在创建 QApplication 之前设置环境变量
    # QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    # QCoreApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)
    is_AA_EnableHighDpiScaling = eval(get_setting_data_from_ini("Config", "高DPI自适应"))
    if is_AA_EnableHighDpiScaling:
        QtCore.QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    app = QtWidgets.QApplication(sys.argv)
    # 防多开
    share = QSharedMemory(APP_NAME)
    share.setKey(APP_NAME)
    if share.attach():
        show_window(APP_NAME)  # 显示窗口
    if share.create(1):
        splash = QSplashScreen(QPixmap(r"./flat/开屏.png"))  # 创建启动界面
        splash.showMessage(
            "加载中......",
            QtCore.Qt.AlignmentFlag.AlignHCenter | QtCore.Qt.AlignmentFlag.AlignBottom,
            QtGui.QColor('green')
        )
        splash.setFont(QFont("微软雅黑", 15))  # 设置字体
        splash.show()  # 显示启动界面

        main_win = Main_window()  # 创建主窗体
        # # 设置窗体样式
        try:
            style_name = "Combinear"
            style_file = r"./flat/{}.qss".format(style_name)
            style_sheet = QSSLoader.read_qss_file(style_file)
            main_win.setStyleSheet(style_sheet)
        except FileNotFoundError:
            pass

        main_win.show()  # 显示窗体，并根据设置检查更新

        splash.finish(main_win)  # 隐藏启动界面
        splash.deleteLater()

        sys.exit(app.exec_())
